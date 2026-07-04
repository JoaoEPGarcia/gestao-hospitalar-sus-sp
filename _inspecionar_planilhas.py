# -*- coding: utf-8 -*-
"""
Script de inspeção das planilhas SIH/SUS e Classificação Hospitais.
Lê SOMENTE metadados (abas, cabeçalhos, primeiras linhas) — não altera nada.
Roda antes do script de análise para validar premissas.
"""

import os
import sys
import io
import glob
import unicodedata
import openpyxl

# Força stdout em UTF-8 para evitar UnicodeEncodeError no Windows (cp1252)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

PASTA = os.path.dirname(os.path.abspath(__file__))

# Colunas canônicas esperadas na aba do estado (referência 2024)
COLUNAS_CANONICAS = [
    "CNES", "NOME FANTASIA", "CO_PROCEDIMENTO", "NO_PROCEDIMENTO",
    "QTDE", "SomaDeDIAS_PERM", "SomaDeUTI_MES_TO", "VALOR TOTAL",
    "DESFECHO", "FINANCIMANTO", "COMPLEX", "Faixa", "SEXO",
    "Cód Grupo", "Nome do Grupo", "Cód Subgrupo", "Nome Subgrupo",
    "Classificação assistencial", "Leitos Internação", "Leitos SUS",
    "UTI ", "UTI SUS", "Total Leitos", "Total Leitos SUS",
    "Diárias Internação Ano", "Diárias UTI Ano", "Porte Hospital",
    "Ocupação Internação", "Ocupação UTI", "Cód IBGE", "Município",
    "Tipo de Hospital", "Especialização",
]

# Mapeamento de variações conhecidas → nome canônico
VARIACOES_CONHECIDAS = {
    "FINANCIAMENTO": "FINANCIMANTO",          # 2021
    "Classificação Assistencial": "Classificação assistencial",  # 2025
    "UTI": "UTI ",                            # sem espaço final (caso exista)
}

COLUNAS_CANONICAS_SET = set(COLUNAS_CANONICAS)

# Colunas canônicas da planilha de classificação
COLUNAS_CLASSIF = [
    "CNES", "Instituição", "Leitos", "Salas Cirúrgicas", "UTI",
    "Cirurgia Torácica", "Neurocirurgia", "Cirurgia Pediátrica",
    "Pontuação", "Classificação", "Hospital?", "cód IBGE", "UF",
]


def normalizar_cabecalho(nome):
    """Aplica variações conhecidas a um nome de coluna, strip de espaços externos."""
    if nome is None:
        return None
    stripped = nome.strip()
    # aplica mapa de variações (mas preserva espaço interno de "UTI ")
    return VARIACOES_CONHECIDAS.get(stripped, nome)  # usa nome original se não há variação


def cabecalhos_aba(ws, max_linhas_header=3):
    """
    Retorna (lista_de_cabecalhos_linha1, primeiras_N_linhas_de_dados).
    Usa a primeira linha como cabeçalho.
    """
    linhas = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        linhas.append(row)
        if i >= max_linhas_header:
            break
    if not linhas:
        return [], []
    header = list(linhas[0])
    dados = linhas[1:]
    return header, dados


def inspecionar_sih(path, ano):
    """Inspeciona um arquivo SIH anual. Retorna dict com resultado da inspeção."""
    resultado = {
        "arquivo": os.path.basename(path),
        "ano": ano,
        "abas": [],
        "aba_estado": None,
        "aba_capital": None,
        "cabecalhos_brutos": [],
        "cabecalhos_normalizados": [],
        "colunas_ausentes": [],
        "colunas_extras": [],
        "colunas_none": 0,
        "linhas_amostra": [],
        "alertas": [],
    }

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    resultado["abas"] = wb.sheetnames

    # Identificar aba do estado (não-Capital)
    aba_estado = None
    aba_capital = None
    for nome_aba in wb.sheetnames:
        if "capital" in nome_aba.lower():
            aba_capital = nome_aba
        else:
            aba_estado = nome_aba

    if aba_estado is None:
        resultado["alertas"].append("ERRO: nenhuma aba não-Capital encontrada!")
        wb.close()
        return resultado

    if aba_capital is None:
        resultado["alertas"].append("AVISO: aba Capital não encontrada (esperava 2 abas).")

    resultado["aba_estado"] = aba_estado
    resultado["aba_capital"] = aba_capital

    ws = wb[aba_estado]
    header_bruto, dados_amostra = cabecalhos_aba(ws, max_linhas_header=5)
    resultado["cabecalhos_brutos"] = header_bruto

    # Contar Nones
    nones = sum(1 for c in header_bruto if c is None)
    resultado["colunas_none"] = nones

    # Normalizar (strip + variações)
    header_norm = []
    for c in header_bruto:
        if c is None:
            header_norm.append(None)
        else:
            header_norm.append(normalizar_cabecalho(c))
    resultado["cabecalhos_normalizados"] = header_norm

    # Colunas reais (sem None)
    reais = [c for c in header_norm if c is not None]
    reais_set = set(reais)

    ausentes = [c for c in COLUNAS_CANONICAS if c not in reais_set]
    extras = [c for c in reais if c not in COLUNAS_CANONICAS_SET]

    resultado["colunas_ausentes"] = ausentes
    resultado["colunas_extras"] = extras

    if ausentes:
        resultado["alertas"].append(f"DIVERGÊNCIA: colunas canônicas ausentes: {ausentes}")
    if extras:
        resultado["alertas"].append(f"INFO: colunas extras (não-canônicas, não-None): {extras}")

    # Guardar amostra de dados (2 linhas)
    resultado["linhas_amostra"] = [list(r) for r in dados_amostra[:2]]

    # Verificar coluna QTDE e Total Leitos para a regra de separação
    idx_qtde = next((i for i, c in enumerate(header_norm) if c == "QTDE"), None)
    idx_tleitos = next((i for i, c in enumerate(header_norm) if c == "Total Leitos"), None)
    resultado["idx_QTDE"] = idx_qtde
    resultado["idx_Total_Leitos"] = idx_tleitos

    if idx_qtde is None:
        resultado["alertas"].append("ERRO: coluna QTDE não encontrada!")
    if idx_tleitos is None:
        resultado["alertas"].append("ERRO: coluna 'Total Leitos' não encontrada!")

    wb.close()
    return resultado


def inspecionar_classificacao(path):
    """Inspeciona a planilha de classificação hospitalar. Retorna dict."""
    resultado = {
        "arquivo": os.path.basename(path),
        "abas": [],
        "aba_alvo": None,
        "cabecalhos_brutos": [],
        "colunas_ausentes": [],
        "colunas_extras": [],
        "n_linhas_amostra": 0,
        "linhas_amostra": [],
        "alertas": [],
    }

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    resultado["abas"] = wb.sheetnames

    ABA_ALVO = "Com Municípios"
    if ABA_ALVO not in wb.sheetnames:
        resultado["alertas"].append(f"ERRO: aba '{ABA_ALVO}' não encontrada. Abas existentes: {wb.sheetnames}")
        wb.close()
        return resultado

    resultado["aba_alvo"] = ABA_ALVO
    ws = wb[ABA_ALVO]
    header_bruto, dados = cabecalhos_aba(ws, max_linhas_header=5)
    resultado["cabecalhos_brutos"] = header_bruto

    reais_set = set(c for c in header_bruto if c is not None)
    esperados_set = set(COLUNAS_CLASSIF)

    resultado["colunas_ausentes"] = [c for c in COLUNAS_CLASSIF if c not in reais_set]
    resultado["colunas_extras"] = [c for c in reais_set if c not in esperados_set]

    if resultado["colunas_ausentes"]:
        resultado["alertas"].append(f"DIVERGÊNCIA: colunas ausentes: {resultado['colunas_ausentes']}")
    if resultado["colunas_extras"]:
        resultado["alertas"].append(f"INFO: colunas extras: {resultado['colunas_extras']}")

    resultado["linhas_amostra"] = [list(r) for r in dados[:3]]

    wb.close()
    return resultado


def imprimir_resultado_sih(r):
    sep = "-" * 70
    print(sep)
    print(f"  ARQUIVO : {r['arquivo']}  (ano={r['ano']})")
    print(f"  Abas    : {r['abas']}")
    print(f"  Aba estado  : {r['aba_estado']}")
    print(f"  Aba capital : {r['aba_capital']}")
    print(f"  Colunas None (padding): {r['colunas_none']}")
    print(f"  Cabeçalhos brutos ({len(r['cabecalhos_brutos'])}):")
    print(f"    {r['cabecalhos_brutos']}")
    print(f"  Cabeçalhos normalizados (sem None) ({len([c for c in r['cabecalhos_normalizados'] if c is not None])}):")
    print(f"    {[c for c in r['cabecalhos_normalizados'] if c is not None]}")
    if r["colunas_ausentes"]:
        print(f"  *** AUSENTES: {r['colunas_ausentes']}")
    if r["colunas_extras"]:
        print(f"  --- EXTRAS  : {r['colunas_extras']}")
    print(f"  idx QTDE={r.get('idx_QTDE')}  idx Total Leitos={r.get('idx_Total_Leitos')}")
    print(f"  Amostra linha 1: {r['linhas_amostra'][0] if r['linhas_amostra'] else 'N/A'}")
    if r["alertas"]:
        for a in r["alertas"]:
            print(f"  !!! {a}")
    print()


def imprimir_resultado_classif(r):
    sep = "=" * 70
    print(sep)
    print(f"  ARQUIVO CLASSIFICAÇÃO : {r['arquivo']}")
    print(f"  Abas  : {r['abas']}")
    print(f"  Aba   : {r['aba_alvo']}")
    print(f"  Cabeçalhos ({len(r['cabecalhos_brutos'])}):")
    print(f"    {r['cabecalhos_brutos']}")
    if r["colunas_ausentes"]:
        print(f"  *** AUSENTES : {r['colunas_ausentes']}")
    if r["colunas_extras"]:
        print(f"  --- EXTRAS   : {r['colunas_extras']}")
    print(f"  Amostra linha 1: {r['linhas_amostra'][0] if r['linhas_amostra'] else 'N/A'}")
    print(f"  Amostra linha 2: {r['linhas_amostra'][1] if len(r['linhas_amostra']) > 1 else 'N/A'}")
    if r["alertas"]:
        for a in r["alertas"]:
            print(f"  !!! {a}")
    print(sep)
    print()


def main():
    print("=" * 70)
    print("INSPEÇÃO DE PLANILHAS SIH/SUS + CLASSIFICAÇÃO")
    print("=" * 70)
    print()

    # --- Planilha de classificação ---
    # Localiza por glob para não depender de normalização NFC/NFD do nome
    padrao_classif = os.path.join(PASTA, "20260303*.xlsx")
    candidatos_classif = glob.glob(padrao_classif)
    if not candidatos_classif:
        print(f"!!! CLASSIFICAÇÃO NÃO ENCONTRADA com padrão '20260303*.xlsx'")
        nome_classif = None
        path_classif = None
    else:
        path_classif = candidatos_classif[0]
        nome_classif = os.path.basename(path_classif)
        r = inspecionar_classificacao(path_classif)
        imprimir_resultado_classif(r)

    # --- Arquivos SIH anuais ---
    # Padrão: nome começa com 8 dígitos, depois " - ", depois 4 dígitos de ano
    import re
    padrao_sih = re.compile(r"^(\d{8})\s*-\s*(\d{4})\.xlsx$", re.IGNORECASE)

    todos_xlsx = sorted(glob.glob(os.path.join(PASTA, "*.xlsx")))
    arquivos_sih = []
    arquivos_fora_padrao = []

    for path in todos_xlsx:
        nome = os.path.basename(path)
        # Compara com normalização NFC para lidar com nomes NFD (macOS/OneDrive)
        if path_classif and os.path.normcase(os.path.abspath(path)) == os.path.normcase(os.path.abspath(path_classif)):
            continue  # já tratado
        m = padrao_sih.match(nome)
        if m:
            ano = int(m.group(2))
            arquivos_sih.append((path, ano, nome))
        else:
            arquivos_fora_padrao.append(nome)

    if arquivos_fora_padrao:
        print(">>> ATENÇÃO: arquivos xlsx FORA DO PADRÃO (não serão processados):")
        for n in arquivos_fora_padrao:
            print(f"     {n}")
        print()

    arquivos_sih.sort(key=lambda x: x[1])
    print(f"Arquivos SIH encontrados: {len(arquivos_sih)}")
    for _, ano, nome in arquivos_sih:
        print(f"  {ano}: {nome}")
    print()

    print("INSPECIONANDO CADA ARQUIVO SIH...")
    print()

    resultados = []
    for path, ano, nome in arquivos_sih:
        r = inspecionar_sih(path, ano)
        imprimir_resultado_sih(r)
        resultados.append(r)

    # Resumo de divergências
    print("=" * 70)
    print("RESUMO DE DIVERGÊNCIAS / ALERTAS")
    print("=" * 70)
    divergencias = [(r["arquivo"], r["alertas"]) for r in resultados if r["alertas"]]
    if not divergencias:
        print("Nenhuma divergência encontrada — todas as premissas validadas.")
    else:
        for arq, alertas in divergencias:
            print(f"\n{arq}:")
            for a in alertas:
                print(f"  {a}")

    print()
    print("Inspeção concluída.")


if __name__ == "__main__":
    main()
