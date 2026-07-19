# -*- coding: utf-8 -*-
"""
_investigacao_etapa1_13jul2026.py
=================================
ETAPA 1 da revisão de 13/07/2026 (reunião João + Priscilla + Alberto) —
INVESTIGAÇÃO SOMENTE-LEITURA. Este script NÃO altera nenhum arquivo do
pipeline: apenas lê o painel bruto, o painel definitivo, as tabelas já
produzidas e as duas planilhas novas trazidas pela equipe, e imprime os
subsídios quantitativos para os itens 1.1 a 1.12 do encaminhamento.

Os itens puramente de código/texto (1.2 rótulo de custo, 1.4 inventário de
gráficos em log, 1.12 inventário de usos em .py/.R/.tex) são cobertos por
greps reprodutíveis fora deste script; aqui entram apenas as partes que
dependem de DADOS.

USO: python _investigacao_etapa1_13jul2026.py
"""

import re
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

import analise_sih as base                      # embrulha stdout em UTF-8
import construir_painel_definitivo as cpd

LARG = 84
PASTA_TAB   = base.PASTA_TABELAS
PASTA_FASE2 = base.PASTA_DADOS / "resultados_fase2" / "tabelas"

CNES_SOROCABA = 2081695
CNES_PEROLA   = 2078287
SWITCHERS     = {2078287, 2081695, 2082225, 2091755, 2750511}

# IPCA anual (variação % dez/dez, IBGE) — MESMA série já usada em
# estimacao.py / analise_exploratoria.py / preparo_fase2.R; replicada aqui
# apenas para leitura (não importa estimacao.py para não carregar statsmodels).
IPCA_ANUAL = {2015: 10.67, 2016: 6.29, 2017: 2.95, 2018: 3.75, 2019: 4.31,
              2020: 4.52, 2021: 10.06, 2022: 5.79, 2023: 4.62, 2024: 4.83,
              2025: 4.26}


def fatores_ipca_2025() -> dict:
    anos = sorted(IPCA_ANUAL)
    acum_ate = {}
    acum = 1.0
    for a in anos:
        acum *= 1 + IPCA_ANUAL[a] / 100
        acum_ate[a] = acum
    total = acum_ate[anos[-1]]
    return {a: total / acum_ate[a] for a in anos}


def titulo(txt: str):
    print("\n" + "=" * LARG)
    print(txt)
    print("=" * LARG)


def sub(txt: str):
    print("\n--- " + txt)


def carregar():
    bruto = cpd.carregar_painel_bruto()
    df = pd.read_csv(cpd.PAINEL_DEFINITIVO_CSV, encoding="utf-8-sig")
    df["cnes"] = pd.to_numeric(df["cnes"], errors="raise").astype("int64")
    df["ano"] = df["ano"].astype(int)
    return bruto, df


def municipio_por_cnes(bruto: pd.DataFrame) -> pd.Series:
    return (bruto[bruto["municipio"].notna()]
            .sort_values("ano").groupby("cnes")["municipio"].last())


# ══════════════════════════════════════════════════════════════════════════════
# ITEM 1.1 — ALTA COMPLEXIDADE NOS FILANTRÓPICOS
# ══════════════════════════════════════════════════════════════════════════════

def item_1_1(bruto: pd.DataFrame, df: pd.DataFrame):
    titulo("ITEM 1.1 — ALTA COMPLEXIDADE: proporção interna × participação estadual")

    sub("Reprodução do número atual (mediana de pct_alta_complex por categoria, "
        "painel definitivo)")
    med = df.groupby("modelo_gestao_proxy")["pct_alta_complex"].median()
    print(med.round(6).to_string())
    print("(comparar com tab_def_por_modelo_gestao.csv: Filantrópico 0.0011)")

    sub("Distribuição interna do Filantrópico (pct_alta_complex, hospital-ano)")
    fil = df[df["modelo_gestao_proxy"] == "Filantrópico"]
    d = fil["pct_alta_complex"].describe(percentiles=[.25, .5, .75, .9, .99])
    print(d.round(5).to_string())
    print(f"fração de hospital-ano com pct_alta_complex = 0: "
          f"{(fil['pct_alta_complex'] == 0).mean():.1%}")

    sub("Participação de cada categoria no TOTAL ESTADUAL de internações de "
        "alta complexidade — painel DEFINITIVO (314 CNES)")
    agg = (df.groupby("modelo_gestao_proxy")
             .agg(qtde_alta=("qtde_alta_complex", "sum"),
                  qtde_tot=("qtde", "sum"),
                  cnes=("cnes", "nunique")))
    agg["prop_interna_agregada"] = agg["qtde_alta"] / agg["qtde_tot"]
    agg["particip_estadual_alta"] = agg["qtde_alta"] / agg["qtde_alta"].sum()
    agg["particip_estadual_qtde"] = agg["qtde_tot"] / agg["qtde_tot"].sum()
    print(agg.round(4).to_string())

    sub("Mesma conta na BASE BRUTA (830 CNES, sem filtros), por class_assistencial")
    br = bruto.copy()
    aggb = (br.groupby("class_assistencial")
              .agg(qtde_alta=("qtde_alta_complex", "sum"),
                   qtde_tot=("qtde", "sum"),
                   cnes=("cnes", "nunique")))
    aggb["prop_interna_agregada"] = aggb["qtde_alta"] / aggb["qtde_tot"]
    aggb["particip_estadual_alta"] = aggb["qtde_alta"] / aggb["qtde_alta"].sum()
    print(aggb.sort_values("particip_estadual_alta", ascending=False)
          .round(4).to_string())

    sub("Indício de bug de junção? CNES filantrópicos (definitivo) com "
        "qtde_alta_complex = 0 em TODOS os anos, mas com produção normal")
    tot = (fil.groupby("cnes")
              .agg(alta_total=("qtde_alta_complex", "sum"),
                   qtde_total=("qtde", "sum"),
                   qtde_min=("qtde", "min")))
    zerados = tot[(tot["alta_total"] == 0) & (tot["qtde_total"] > 0)]
    print(f"{len(zerados)} de {len(tot)} CNES filantrópicos sem NENHUMA "
          f"internação de alta complexidade em 11 anos.")
    conc = tot.sort_values("alta_total", ascending=False)
    conc["particip_grupo"] = conc["alta_total"] / conc["alta_total"].sum()
    print("\nTop 10 filantrópicos por volume de alta complexidade "
          "(concentração dentro do grupo):")
    print(conc.head(10).round(4).to_string())
    print(f"Os 10 maiores concentram "
          f"{conc['particip_grupo'].head(10).sum():.1%} da alta complexidade "
          f"do grupo Filantrópico.")


# ══════════════════════════════════════════════════════════════════════════════
# ITEM 1.3 — OCUPAÇÃO SEM 2020–2021
# ══════════════════════════════════════════════════════════════════════════════

def item_1_3(df: pd.DataFrame):
    titulo("ITEM 1.3 — OCUPAÇÃO EXCLUINDO 2020–2021 (prévia, nada salvo)")

    pand = df["ano"].isin([2020, 2021])
    for c in ["ocupacao_internacao", "ocupacao_uti"]:
        n_total = df[c].notna().sum()
        n_rem = df.loc[pand, c].notna().sum()
        print(f"{c}: {n_total} obs não-nulas no total; {n_rem} seriam removidas "
              f"ao excluir 2020–2021 ({n_rem / n_total:.1%})")
    sub("Observações removidas por categoria (linhas 2020–2021 com ocupação "
        "não-nula)")
    print(df[pand].groupby("modelo_gestao_proxy")[
        ["ocupacao_internacao", "ocupacao_uti"]].count().to_string())

    sub("PRÉVIA da tabela comparativa (mediana | média), por categoria")
    linhas = []
    for versao, dfx in [("2015–2025 completo", df),
                        ("excluindo 2020–2021", df[~pand])]:
        g = dfx.groupby("modelo_gestao_proxy")[
            ["ocupacao_internacao", "ocupacao_uti"]]
        t = pd.concat([g.median().add_suffix("_mediana"),
                       g.mean().add_suffix("_media")], axis=1)
        t["versao"] = versao
        linhas.append(t.reset_index())
    tab = (pd.concat(linhas, ignore_index=True)
           .set_index(["modelo_gestao_proxy", "versao"]).sort_index())
    print(tab.round(2).to_string())

    sub("Agregado (todas as categorias)")
    for versao, dfx in [("2015–2025 completo", df),
                        ("excluindo 2020–2021", df[~pand])]:
        m = dfx[["ocupacao_internacao", "ocupacao_uti"]]
        print(f"{versao}: ocup_int mediana {m['ocupacao_internacao'].median():.2f} "
              f"média {m['ocupacao_internacao'].mean():.2f} | "
              f"ocup_uti mediana {m['ocupacao_uti'].median():.2f} "
              f"média {m['ocupacao_uti'].mean():.2f}")


# ══════════════════════════════════════════════════════════════════════════════
# ITEM 1.4 — TMP ACIMA DE 120 DIAS (subsídio para a decisão de escala)
# ══════════════════════════════════════════════════════════════════════════════

def item_1_4(df: pd.DataFrame, bruto: pd.DataFrame):
    titulo("ITEM 1.4 — TMP: quantas observações excedem 120 dias?")

    mun = municipio_por_cnes(bruto)
    for col in ["tmp", "tmp_com_covid"]:
        x = df[col].dropna()
        print(f"{col}: máx {x.max():.1f} dias | >120: {(x > 120).sum()} obs | "
              f">30,5 (teto de longa permanência já usado na estimação): "
              f"{(x > 30.5).sum()} obs | p99 {x.quantile(.99):.1f}")
    alto = df[df["tmp"] > 120][["cnes", "ano", "tmp", "modelo_gestao_proxy"]]
    if len(alto):
        alto = alto.copy()
        alto["municipio"] = alto["cnes"].map(mun)
        print("\nHospital-ano com tmp > 120 dias (painel definitivo):")
        print(alto.to_string(index=False))
    else:
        print("\nNenhuma observação com tmp > 120 no painel definitivo.")
    print("\nCNES com mediana de tmp > 20 dias (grupo 'longa permanência' da "
          "exploratória):")
    med = df[df["tmp"] > 0].groupby("cnes")["tmp"].median()
    lp = med[med > 20].sort_values(ascending=False)
    lp = pd.DataFrame({"tmp_mediano": lp.round(1),
                       "municipio": [mun.get(c, "") for c in lp.index]})
    print(lp.to_string())


# ══════════════════════════════════════════════════════════════════════════════
# ITEM 1.5 — VARIÁVEIS MONETÁRIAS (inventário; correção IPCA já existente)
# ══════════════════════════════════════════════════════════════════════════════

def item_1_5(df: pd.DataFrame):
    titulo("ITEM 1.5 — VARIÁVEIS MONETÁRIAS NO PAINEL DEFINITIVO")

    monet = [c for c in df.columns
             if c.startswith("valor") or c.startswith("custo")]
    print("Colunas monetárias presentes (todas em R$ CORRENTES do ano):")
    for c in monet:
        s = df[c].dropna()
        print(f"  {c:26} n={len(s):5}  mediana R$ {s.median():>12,.2f}")
    sub("Verificação de que são valores correntes: mediana de custo_saida por ano")
    print(df.groupby("ano")["custo_saida"].median().round(2).to_string())
    print("\nNOTA: a correção IPCA JÁ EXISTE no pipeline de estimação "
          "(custo_real = custo_saida × fator IPCA dez/dez IBGE, base 2025) em "
          "estimacao.py, analise_exploratoria.py (fig_ae_07/tab_ae_custo_real_ano) "
          "e resultados_fase2/preparo_fase2.R. O que NÃO é deflacionado hoje: "
          "tabelas/figuras descritivas (tab_def_*, tab_pospatch_*, figD*, "
          "fig_ae_04/05/06/08 e as tabelas por categoria).")
    fat = fatores_ipca_2025()
    print("\nFatores IPCA para preços de 2025 (série já adotada):")
    print(pd.Series(fat).round(4).to_string())


# ══════════════════════════════════════════════════════════════════════════════
# ITEM 1.6 — SOROCABA (2081695) E PÉROLA BYINGTON (2078287)
# ══════════════════════════════════════════════════════════════════════════════

def item_1_6(df: pd.DataFrame):
    titulo("ITEM 1.6 — TRAJETÓRIA ANO A ANO: Sorocaba (2081695) × "
           "Pérola Byington (2078287)")

    fat = fatores_ipca_2025()
    cols = ["ano", "modelo_gestao_proxy", "faixa_barcelona",
            "complexidade_estrutural", "mort_all", "tmp", "custo_saida",
            "qtde_sem_covid", "pct_alta_complex",
            "ocupacao_internacao", "ocupacao_uti"]
    for cnes, rotulo in [(CNES_SOROCABA, "Conjunto Hospitalar Sorocaba "
                                          "(conversão nov/2018 → OSS de 2019 em diante)"),
                         (CNES_PEROLA, "Pérola Byington (conversão em 2023)")]:
        sub(f"CNES {cnes} — {rotulo}")
        t = df.loc[df["cnes"] == cnes, cols].sort_values("ano").copy()
        t["custo_real_2025"] = t["custo_saida"] * t["ano"].map(fat)
        t["pct_alta_complex"] = (100 * t["pct_alta_complex"]).round(2)
        print(t.round(3).to_string(index=False))

    print("\nNOTA METODOLÓGICA IMPORTANTE: complexidade_estrutural e "
          "faixa_barcelona são FIXAS por CNES (vêm da planilha única de "
          "classificação Barcelona) — por construção NÃO podem mostrar queda "
          "pós-conversão. Os únicos sinais de composição que VARIAM no tempo "
          "no painel são pct_alta_complex, volume (qtde) e ocupação.")


# ══════════════════════════════════════════════════════════════════════════════
# ITEM 1.7 — ORIGEM E IC DA ESTIMATIVA DE −6,7% NO CUSTO/FATURAMENTO
# ══════════════════════════════════════════════════════════════════════════════

def item_1_7():
    titulo("ITEM 1.7 — ESTIMATIVA DE −6,7% (transição OSS, custo/faturamento)")

    fe = pd.read_csv(PASTA_TAB / "tab_est_custo_fe.csv", encoding="utf-8-sig",
                     index_col=0)
    b = fe.loc["OSS", "coeficiente"]
    ep = fe.loc["OSS", "ep_cluster"]
    lo, hi = b - 1.96 * ep, b + 1.96 * ep
    print("Fonte primária: tab_est_custo_fe.csv (estimacao.py — log custo_real, "
          "efeitos fixos de hospital e ano, EP agrupado por CNES):")
    print(f"  coef OSS = {b:.5f} log-pontos | EP = {ep:.5f} | "
          f"p = {fe.loc['OSS', 'p_valor']:.4f}")
    print(f"  efeito % = {100 * (np.exp(b) - 1):.2f}%  "
          f"IC95% = [{100 * (np.exp(lo) - 1):.1f}%, {100 * (np.exp(hi) - 1):.1f}%]")

    sub("Réplicas do mesmo número em outros artefatos")
    ro = pd.read_csv(PASTA_TAB / "tab_est_resumo_oss.csv", encoding="utf-8-sig")
    print("tab_est_resumo_oss.csv (linha custo):")
    print(ro[ro["modelo"].str.contains("Custo")].to_string(index=False))
    fr = pd.read_csv(PASTA_FASE2 / "tabR_frente1_variantes.csv")
    fx = fr[fr["modelo"] == "custo"].copy()
    fx["efeito_pct"] = 100 * (np.exp(fx["coef_oss"]) - 1)
    fx["ic_lo_pct"] = 100 * (np.exp(fx["ic_lo"]) - 1)
    fx["ic_hi_pct"] = 100 * (np.exp(fx["ic_hi"]) - 1)
    print("\ntabR_frente1_variantes.csv (fase 2 — CRE/Mundlak, modelo custo):")
    print(fx[["variante", "n", "coef_oss", "ep", "efeito_pct",
              "ic_lo_pct", "ic_hi_pct"]].round(3).to_string(index=False))
    wb = pd.read_csv(PASTA_TAB / "tab_est_wild_bootstrap.csv",
                     encoding="utf-8-sig")
    print("\ntab_est_wild_bootstrap.csv (inferência robusta, 5 clusters tratados):")
    print(wb.to_string(index=False))
    bay = PASTA_FASE2 / "tabB_comparacao_freq_bayes.csv"
    if bay.exists():
        print("\ntabB_comparacao_freq_bayes.csv (linha custo, se houver):")
        t = pd.read_csv(bay)
        col0 = t.columns[0]
        print(t[t[col0].astype(str).str.contains("usto", na=False)]
              .to_string(index=False))


# ══════════════════════════════════════════════════════════════════════════════
# ITEM 1.8 — OCUPAÇÃO DE UTI: MÉTODO ATUAL × PROXY "BARCELONA"
# ══════════════════════════════════════════════════════════════════════════════

def _inspecionar_xlsx(path: Path, max_linhas: int = 4):
    print(f"\nArquivo: {path.name}")
    if not path.exists():
        print("  (não encontrado)")
        return
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for aba in wb.sheetnames:
        ws = wb[aba]
        print(f"  Aba '{aba}' ({ws.max_row}x{ws.max_column}):")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_linhas:
                break
            vals = ["" if v is None else str(v)[:28] for v in row[:12]]
            print("    " + " | ".join(vals))
    wb.close()


def item_1_8(df: pd.DataFrame):
    titulo("ITEM 1.8 — OCUPAÇÃO DE UTI: o que existe hoje nos dados (sem código "
           "novo, só mapeamento)")

    cl = pd.read_csv(PASTA_TAB / "tab_classificacao_hospitais.csv",
                     encoding="utf-8-sig")
    print("Colunas da planilha de classificação Barcelona "
          "(tab_classificacao_hospitais.csv):")
    print("  " + ", ".join(cl.columns))
    print("→ NÃO há campo de ocupação (nem de UTI-dia) na classificação "
          "Barcelona; o único insumo de UTI é o Nº DE LEITOS de UTI (peso 4 na "
          "pontuação).")

    sub("ocupacao_uti atual (dias-UTI ÷ leitos-UTI×365, do resumo SIH) — "
        "mediana por ano")
    print(df.groupby("ano")["ocupacao_uti"].median().round(1).to_string())

    sub("Inspeção das duas planilhas novas trazidas pela equipe (podem conter "
        "material da reunião)")
    _inspecionar_xlsx(base.PASTA_DADOS / "Classificacao_Assistencial_UFSCar.xlsx")
    _inspecionar_xlsx(base.PASTA_DADOS / "Hospitais_por_Categoria_Painel314.xlsx")


# ══════════════════════════════════════════════════════════════════════════════
# ITEM 1.10 — CANDIDATOS PEDIÁTRICOS / ONCO-PEDIÁTRICOS
# ══════════════════════════════════════════════════════════════════════════════

PAT_ESPEC = re.compile(r"PEDIATR|INFANTIL|CRIANCA")
PAT_NOME = re.compile(
    r"INFANTIL|PEDIATR|CRIANCA|BOLDRINI|DARCY VARGAS|GRAACC|ITACI|"
    r"ONCOLOGIA INFANTIL|MARTAGAO|PEQUENO PRINCIPE|SABARA")
PAT_MATERNO = re.compile(r"MATERNO")


def item_1_10(bruto: pd.DataFrame, df: pd.DataFrame):
    titulo("ITEM 1.10 — CANDIDATOS A HOSPITAL PEDIÁTRICO/ONCO-PEDIÁTRICO "
           "(levantamento para revisão manual — NADA é excluído)")

    def_cnes = set(df["cnes"].unique())

    sub("Valores DISTINTOS de tipo_hospital no painel definitivo")
    print(df["tipo_hospital"].fillna("(nulo)").value_counts().to_string())
    sub("Valores DISTINTOS de especializacao no painel definitivo")
    print(df["especializacao"].fillna("(nulo)").value_counts().to_string())

    # ── Sinal 1: especialização declarada em qualquer ano (base bruta) ───────
    esp = bruto[["cnes", "especializacao"]].dropna(subset=["especializacao"]).copy()
    esp["norm"] = esp["especializacao"].map(cpd.normalizar_texto)
    hit_esp = esp[esp["norm"].str.contains(PAT_ESPEC, regex=True)]
    cnes_esp = set(hit_esp["cnes"])
    print(f"\n[Sinal A] especializacao com PEDIATR/INFANTIL/CRIANCA "
          f"(base bruta, qualquer ano): {sorted(hit_esp['especializacao'].unique())} "
          f"→ {len(cnes_esp)} CNES")

    # ── Sinal 2: nome_fantasia (só preenchido em 2020-2021) ─────────────────
    nm = bruto[["cnes", "nome_fantasia"]].dropna(subset=["nome_fantasia"]).copy()
    nm["norm"] = nm["nome_fantasia"].map(cpd.normalizar_texto)
    hit_nome = nm[nm["norm"].str.contains(PAT_NOME, regex=True)]
    cnes_nome = set(hit_nome["cnes"])
    print(f"[Sinal B] nome_fantasia (2020-2021) com termos pediátricos: "
          f"{len(cnes_nome)} CNES")

    # ── Sinal 3: Instituição na classificação Barcelona (nomes completos) ────
    cl = pd.read_csv(PASTA_TAB / "tab_classificacao_hospitais.csv",
                     encoding="utf-8-sig")
    cl["cnes"] = pd.to_numeric(cl["CNES"], errors="coerce")
    cl = cl.dropna(subset=["cnes"])
    cl["cnes"] = cl["cnes"].astype("int64")
    cl["norm"] = cl["Instituição"].map(cpd.normalizar_texto)
    hit_inst = cl[cl["norm"].str.contains(PAT_NOME, regex=True)]
    cnes_inst = set(hit_inst["cnes"])
    print(f"[Sinal C] 'Instituição' da classificação Barcelona com termos "
          f"pediátricos: {len(cnes_inst)} CNES")

    # ── Sinal D (revisão manual, não exclusão): materno-infantil ────────────
    cnes_mat = set(nm[nm["norm"].str.contains(PAT_MATERNO, regex=True)]["cnes"]) | \
               set(cl[cl["norm"].str.contains(PAT_MATERNO, regex=True)]["cnes"])
    print(f"[Sinal D] 'MATERNO' em nome/Instituição (candidato a revisão, "
          f"maternidades permanecem): {len(cnes_mat)} CNES")

    uniao = sorted((cnes_esp | cnes_nome | cnes_inst | cnes_mat))
    nomes = cpd.nome_referencia_por_cnes(bruto)
    mun = municipio_por_cnes(bruto)
    espec_modal = cpd.valor_modal_por_cnes(bruto, "especializacao")
    tipo_modal = cpd.valor_modal_por_cnes(bruto, "tipo_hospital")
    inst = cl.set_index("cnes")["Instituição"]
    anos_def = df.groupby("cnes")["ano"].nunique()
    porte = df.sort_values("ano").groupby("cnes")["porte_hospital"].last()
    proxy = df.sort_values("ano").groupby("cnes")["modelo_gestao_proxy"].last()

    linhas = []
    for c in uniao:
        vias = []
        if c in cnes_esp:  vias.append("espec")
        if c in cnes_nome: vias.append("nome")
        if c in cnes_inst: vias.append("instituicao")
        if c in cnes_mat:  vias.append("materno(rev.manual)")
        linhas.append({
            "cnes": c,
            "nome": (nomes.get(c) or inst.get(c, ""))[:42],
            "municipio": str(mun.get(c, ""))[:20],
            "no_def": "SIM" if c in def_cnes else "não",
            "anos_def": int(anos_def.get(c, 0)),
            "porte": porte.get(c, ""),
            "modelo_gestao": proxy.get(c, ""),
            "tipo_modal": str(tipo_modal.get(c, ""))[:26],
            "espec_modal": str(espec_modal.get(c, ""))[:22],
            "switcher": "SIM" if c in SWITCHERS else "",
            "via": "+".join(vias),
        })
    t = pd.DataFrame(linhas)
    sub(f"União dos sinais: {len(t)} CNES candidatos "
        f"({(t['no_def'] == 'SIM').sum()} dentro do painel definitivo de 314)")
    with pd.option_context("display.width", 400, "display.max_columns", 20,
                           "display.max_rows", 300):
        print(t.to_string(index=False))
    dentro = t[t["no_def"] == "SIM"]
    alerta = dentro[(dentro["switcher"] == "SIM")
                    | (dentro["modelo_gestao"] == "Privado")]
    sub("Cruzamentos de atenção (switcher ou grupo Privado)")
    print(alerta.to_string(index=False) if len(alerta) else "(nenhum)")
    print("\n>>> NADA foi excluído: lista para validação linha a linha "
          "(Etapa 2, item 7 da ordem sugerida).")


# ══════════════════════════════════════════════════════════════════════════════
# ITEM 1.11 — HOSPITAL GUILHERME ALVARO (SANTOS)
# ══════════════════════════════════════════════════════════════════════════════

def item_1_11(bruto: pd.DataFrame, df: pd.DataFrame):
    titulo("ITEM 1.11 — HOSPITAL GUILHERME ALVARO (Santos): rótulo OSS × Direta")

    nm = bruto[["cnes", "ano", "nome_fantasia", "municipio"]].copy()
    nm["norm"] = nm["nome_fantasia"].map(cpd.normalizar_texto)
    hit = nm[nm["norm"].str.contains("GUILHERME ALVARO", na=False)]
    cl = pd.read_csv(PASTA_TAB / "tab_classificacao_hospitais.csv",
                     encoding="utf-8-sig")
    cl["norm"] = cl["Instituição"].map(cpd.normalizar_texto)
    hit_cl = cl[cl["norm"].str.contains("GUILHERME ALVARO", na=False)]
    cnes_set = set(hit["cnes"]) | set(pd.to_numeric(
        hit_cl["CNES"], errors="coerce").dropna().astype("int64"))
    print(f"CNES localizados por nome: {sorted(cnes_set)}")
    if hit_cl.shape[0]:
        print("Na classificação Barcelona:")
        print(hit_cl[["CNES", "Instituição", "Classificação",
                      "Pontuação"]].to_string(index=False))

    for c in sorted(cnes_set):
        sub(f"CNES {c} — histórico na BASE BRUTA")
        s = (bruto[bruto["cnes"] == c]
             [["ano", "class_assistencial", "qtde", "municipio"]]
             .sort_values("ano"))
        s["producao"] = np.where(s["qtde"] > 0, "com produção", "SEM produção")
        print(s[["ano", "class_assistencial", "producao", "municipio"]]
              .to_string(index=False))
        vals = sorted(s["class_assistencial"].dropna().unique())
        print(f"Valores distintos na série: {vals} → "
              f"{'rótulo ESTÁVEL (erro seria de origem, todos os anos)' if len(vals) == 1 else 'MUDANÇA de rótulo ao longo dos anos'}")
        no_def = df[df["cnes"] == c].sort_values("ano")
        if no_def.empty:
            print("No painel DEFINITIVO: AUSENTE (removido por filtro).")
        else:
            print(f"No painel DEFINITIVO ({len(no_def)}/11 anos) — "
                  f"modelo_gestao_proxy por ano:")
            print("  " + "; ".join(f"{int(r.ano)}={r.modelo_gestao_proxy}"
                                   for r in no_def.itertuples()))
        print(f"É um dos 5 switchers Direta→OSS documentados? "
              f"{'SIM' if c in SWITCHERS else 'NÃO'}")


# ══════════════════════════════════════════════════════════════════════════════
# ITEM 1.12 — INVENTÁRIO DE ARTEFATOS COM mort_sem_excl / complexidade_pond_mort
# ══════════════════════════════════════════════════════════════════════════════

def item_1_12():
    titulo("ITEM 1.12 — TABELAS/RESULTADOS que contêm mort_sem_excl ou "
           "complexidade_pond_mort (varredura de cabeçalhos e conteúdo)")

    alvos = ["mort_sem_excl", "complexidade_pond_mort"]
    pastas = [PASTA_TAB, PASTA_FASE2]
    achados = []
    for pasta in pastas:
        for p in sorted(pasta.glob("*.csv")):
            if p.stat().st_size > 8_000_000:
                continue
            try:
                texto = p.read_text(encoding="utf-8-sig", errors="replace")
            except Exception:
                continue
            header = texto.splitlines()[0] if texto else ""
            for alvo in alvos:
                if alvo in texto:
                    onde = "COLUNA" if alvo in header else "conteúdo"
                    achados.append({"arquivo": str(p.relative_to(base.PASTA_DADOS)),
                                    "alvo": alvo, "onde": onde})
    t = pd.DataFrame(achados)
    if len(t):
        with pd.option_context("display.max_colwidth", 70,
                               "display.max_rows", 200):
            print(t.to_string(index=False))
    else:
        print("(nenhum CSV com os termos)")
    print("\nOBS: o inventário de CÓDIGO/TEXTO (.py/.R/.tex/.md) foi feito por "
          "grep e está no relatório da Etapa 1.")


def main():
    print("=" * LARG)
    print("INVESTIGAÇÃO ETAPA 1 — REVISÃO DE 13/07/2026 — SOMENTE LEITURA")
    print("=" * LARG)
    base.configurar_diretorios()
    bruto, df = carregar()
    print(f"[CARGA] bruto: {bruto['cnes'].nunique()} CNES / {len(bruto)} linhas | "
          f"definitivo: {df['cnes'].nunique()} CNES / {len(df)} linhas")

    item_1_1(bruto, df)
    item_1_3(df)
    item_1_4(df, bruto)
    item_1_5(df)
    item_1_6(df)
    item_1_7()
    item_1_8(df)
    item_1_10(bruto, df)
    item_1_11(bruto, df)
    item_1_12()

    print("\n" + "=" * LARG)
    print("FIM — nenhum arquivo do pipeline foi alterado.")
    print("=" * LARG)


if __name__ == "__main__":
    main()
