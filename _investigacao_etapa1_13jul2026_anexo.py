# -*- coding: utf-8 -*-
"""
_investigacao_etapa1_13jul2026_anexo.py
=======================================
Anexo da investigação da ETAPA 1 (13/07/2026) — SOMENTE LEITURA.

Complementa _investigacao_etapa1_13jul2026.py:
  A) Conteúdo INTEGRAL das duas planilhas novas trazidas pela equipe
     (Classificacao_Assistencial_UFSCar.xlsx e
      Hospitais_por_Categoria_Painel314.xlsx) — relevantes aos itens 1.9/1.10.
  B) Tabelas de modelagem que reportam a mortalidade alternativa sob o rótulo
     "ajustada" (tab_est_mort_zib_ajustada.csv, tabB_sensibilidade_mortalidade)
     e a comparação de ajuste de output do DEA (tabE_spearman_ajuste_output),
     para fechar o inventário do item 1.12.

USO: python _investigacao_etapa1_13jul2026_anexo.py
"""

from pathlib import Path

import openpyxl
import pandas as pd

import analise_sih as base  # embrulha stdout em UTF-8

PASTA_TAB   = base.PASTA_TABELAS
PASTA_FASE2 = base.PASTA_DADOS / "resultados_fase2" / "tabelas"


def dump_xlsx(path: Path):
    print("\n" + "=" * 84)
    print(f"CONTEÚDO INTEGRAL — {path.name}")
    print("=" * 84)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for aba in wb.sheetnames:
        ws = wb[aba]
        print(f"\n--- Aba '{aba}' ({ws.max_row}x{ws.max_column})")
        for row in ws.iter_rows(values_only=True):
            vals = ["" if v is None else str(v) for v in row]
            if any(vals):
                print("  " + " | ".join(vals))
    wb.close()


def dump_csv(path: Path, max_linhas: int = 40):
    print("\n" + "=" * 84)
    print(f"CSV — {path}")
    print("=" * 84)
    if not path.exists():
        print("  (não existe)")
        return
    t = pd.read_csv(path, encoding="utf-8-sig")
    with pd.option_context("display.width", 300, "display.max_columns", 30):
        print(t.head(max_linhas).to_string(index=False))


def main():
    dump_xlsx(base.PASTA_DADOS / "Classificacao_Assistencial_UFSCar.xlsx")
    dump_xlsx(base.PASTA_DADOS / "Hospitais_por_Categoria_Painel314.xlsx")
    dump_csv(PASTA_TAB / "tab_est_mort_zib_ajustada.csv")
    dump_csv(PASTA_FASE2 / "tabB_sensibilidade_mortalidade.csv")
    dump_csv(PASTA_FASE2 / "tabE_spearman_ajuste_output.csv")
    dump_csv(PASTA_TAB / "tab_covid_com_sem.csv")
    print("\nFIM — somente leitura.")


if __name__ == "__main__":
    main()
