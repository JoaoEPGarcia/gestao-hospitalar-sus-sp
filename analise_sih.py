# -*- coding: utf-8 -*-
"""
analise_sih.py
==============
Análise Exploratória SIH/SUS — Hospitais SUS, Estado de São Paulo, 2015-2025.

INSTRUÇÕES:
    pip install pandas openpyxl matplotlib seaborn numpy pyarrow
    python analise_sih.py

Na 2ª execução, o painel parquet é reutilizado automaticamente (evita reprocessamento).
Para forçar reprocessamento completo: apague ./analises/painel_hospital_ano.parquet

DIVERGÊNCIAS TRATADAS NESTE SCRIPT:
  D1. Strip aplicado em TODOS os cabeçalhos de TODAS as planilhas.
      "UTI " (com espaço) vira "UTI"; o nome canônico foi atualizado.
  D2. Arquivo 2025 tem DUAS colunas que mapeiam para o nome canônico
      "Classificação assistencial": posição 17 (cabeçalho "Classificação
      assistencial", valor CONSTANTE — não varia por CNES; INCORRETA) e
      posição ~38 (cabeçalho "Classificação Assistencial", normalizada via
      MAPA_VARIACOES; a CORRETA). A posição ~38 NÃO é coluna por linha: é
      TABELA DE LOOKUP embutida nas primeiras ~637 linhas de produção
      (pares CNES na pos. 37 → rótulo na pos. 38); na linha-resumo vem
      None. Em 2025 a ÚLTIMA ocorrência vence no índice de colunas
      (preferir_ultima={"Classificação assistencial"}) e o rótulo do
      resumo é atribuído por CNES a partir do lookup.
  D3. Planilha de classificação: "UF " → "UF" após strip.
  D4. Arquivo "2015 - AIHS - SP (1).xlsx" removido da pasta pelo usuário.
      Regex de seleção captura apenas padrão "20260... - AAAA.xlsx".
"""

import io
import os
import re
import sys
import unicodedata
import warnings
from pathlib import Path

import matplotlib
matplotlib.use("Agg")   # backend sem janela; grava figuras em disco
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import seaborn as sns

# Força UTF-8 no terminal Windows (evita UnicodeEncodeError com cp1252)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ══════════════════════════════════════════════════════════════════════════════
# A. CAMINHOS
# ══════════════════════════════════════════════════════════════════════════════

PASTA_DADOS    = Path(__file__).parent
PASTA_ANALISES = PASTA_DADOS / "analises"
PASTA_FIGURAS  = PASTA_ANALISES / "figuras"
PASTA_TABELAS  = PASTA_ANALISES / "tabelas"
PAINEL_PARQUET = PASTA_ANALISES / "painel_hospital_ano.parquet"
PAINEL_CSV     = PASTA_ANALISES / "painel_hospital_ano.csv"


# ══════════════════════════════════════════════════════════════════════════════
# B. CONSTANTES
# ══════════════════════════════════════════════════════════════════════════════

# 33 colunas canônicas — após strip universal, "UTI " torna-se "UTI"
COLUNAS_CANONICAS = [
    "CNES", "NOME FANTASIA", "CO_PROCEDIMENTO", "NO_PROCEDIMENTO",
    "QTDE", "SomaDeDIAS_PERM", "SomaDeUTI_MES_TO", "VALOR TOTAL",
    "DESFECHO", "FINANCIMANTO", "COMPLEX", "Faixa", "SEXO",
    "Cód Grupo", "Nome do Grupo", "Cód Subgrupo", "Nome Subgrupo",
    "Classificação assistencial", "Leitos Internação", "Leitos SUS",
    "UTI", "UTI SUS", "Total Leitos", "Total Leitos SUS",
    "Diárias Internação Ano", "Diárias UTI Ano", "Porte Hospital",
    "Ocupação Internação", "Ocupação UTI", "Cód IBGE", "Município",
    "Tipo de Hospital", "Especialização",
]
COLUNAS_CANONICAS_SET = set(COLUNAS_CANONICAS)

# Variações de nome (pós-strip) → nome canônico
# "UTI " → "UTI" já resolvido pelo strip universal; não precisa entrar aqui.
MAPA_VARIACOES = {
    "FINANCIAMENTO":              "FINANCIMANTO",                 # 2021
    "Classificação Assistencial": "Classificação assistencial",   # 2025 pos~38 — a CORRETA (pos 17 é constante)
}

# Colunas esperadas na planilha de classificação (após strip)
COLUNAS_CLASSIF = [
    "CNES", "Instituição", "Leitos", "Salas Cirúrgicas", "UTI",
    "Cirurgia Torácica", "Neurocirurgia", "Cirurgia Pediátrica",
    "Pontuação", "Classificação", "Hospital?", "cód IBGE", "UF",
]

# String canônica "Óbito" em NFC — usada como prefixo em todos os testes de DESFECHO.
# Necessário mesmo que o arquivo não tenha NFD (proteção para anos futuros).
OBITO_NFC = unicodedata.normalize("NFC", "Óbito")

# Mapeamento explícito de TODOS os valores de DESFECHO que começam com "Óbito"
# encontrados em 2024 (6 categorias; verificar se outros anos adicionam novas).
#
# inclui_B = True  → conta em mort_sem_excl (versão B)
# inclui_B = False → excluído de mort_sem_excl (eventos obstétricos/perinatais)
#
# A DECISÃO CLÍNICA DE CADA LINHA É DA PESQUISADORA — edite à vontade.
# Se surgir um valor desconhecido em algum ano, será logado e incluído por default
# (comportamento conservador — ver eh_obito_versao_b).
DESFECHO_OBITO_VERSAO_B: dict[str, bool] = {
    # ── Mortalidade geral — inequivocamente incluídas em ambas as versões ────
    "Óbito com DO fornecida pelo médico assistente": True,
    "Óbito com DO fornecida pelo SVO":               True,
    "Óbito com DO fornecida pelo IML":               True,
    # ── Óbitos obstétricos — decisão clínica (False = exclui da versão B) ───
    # A mãe morreu; o recém-nascido teve alta vivo.
    "Óbito da mãe/puérpera e alta do recém-nascido":      False,
    # A mãe morreu; o recém-nascido permanece internado.
    "Óbito da mãe/puérpera e permanência recém-nascido":  False,
    # Gestante e concepto morreram (evento obstétrico duplo).
    "Óbito da gestante e do concepto":                     False,
}

# Anos com procedimentos COVID (código 999 em Cód Grupo ou Cód Subgrupo)
ANOS_COVID = {2020, 2021}

# item 1.13 (cruzamento óbito×complexidade) — DESATIVADO por decisão de João
# em 15/07/2026: reversão à definição ORIGINAL de mortalidade (mort_all e
# mort_sem_excl como indicadores paralelos, sem estratificação por
# complexidade). O código permanece no lugar para eventual reativação pela
# equipe: mudar para True e apagar os caches (painel_hospital_ano.* e
# covid_numeradores_2020_2021.csv). Flag ÚNICA — construir_painel_definitivo
# a referencia via base.ITEM_113_MORT_ESTRATIFICADA.
ITEM_113_MORT_ESTRATIFICADA = False

# Cinco indicadores principais
INDICADORES = ["mort_all", "mort_sem_excl", "tmp", "custo_saida", "pct_alta_complex"]
ROTULOS = {
    "mort_all":           "Mortalidade (todos os óbitos)",
    "mort_sem_excl":      "Mortalidade (exc. fetal/materno/neonatal)",
    "tmp":                "Tempo médio de permanência (dias)",
    "custo_saida":        "Faturamento por saída (SIH, R$)",
    "pct_alta_complex":   "% Alta complexidade",
    "mort_alta_complex":  "Mortalidade — internações de alta complexidade",
    "mort_baixa_complex": "Mortalidade — demais internações",
    "ocupacao_internacao":"Ocupação internação",
    "ocupacao_uti":       "Ocupação UTI",
}

sns.set_theme(style="whitegrid", palette="muted", font_scale=0.9)


# ══════════════════════════════════════════════════════════════════════════════
# C. UTILITÁRIOS DE INGESTÃO
# ══════════════════════════════════════════════════════════════════════════════

def configurar_diretorios():
    """Cria pastas de saída se ainda não existirem."""
    for d in [PASTA_ANALISES, PASTA_FIGURAS, PASTA_TABELAS]:
        d.mkdir(parents=True, exist_ok=True)
    print(f"[DIR] Saídas em: {PASTA_ANALISES}")


def localizar_arquivos(pasta: Path):
    """
    Localiza arquivos SIH (padrão 8dígitos - 4dígitos.xlsx) e a planilha
    de classificação (começa com "20260303"). Arquivos fora do padrão são
    reportados e ignorados.

    Retorna:
        arquivos_sih : lista de (Path, int_ano) ordenada por ano
        path_classif : Path ou None
        fora_padrao  : lista de nomes ignorados
    """
    padrao = re.compile(r"^(\d{8})\s*-\s*(\d{4})\.xlsx$", re.IGNORECASE)
    arquivos_sih, fora_padrao = [], []
    path_classif = None

    for p in sorted(pasta.glob("*.xlsx")):
        m = padrao.match(p.name)
        if m:
            arquivos_sih.append((p, int(m.group(2))))
        elif p.name.startswith("20260303"):
            path_classif = p
        else:
            fora_padrao.append(p.name)

    arquivos_sih.sort(key=lambda x: x[1])

    if fora_padrao:
        print("[AVISO] Arquivos fora do padrão — NÃO processados:")
        for n in fora_padrao:
            print(f"          {n}")

    return arquivos_sih, path_classif, fora_padrao


def strip_col(nome):
    """Strip de espaços em cabeçalho; retorna None intacto."""
    return nome.strip() if isinstance(nome, str) else nome


def normalizar_col(nome: str) -> str:
    """Aplica MAPA_VARIACOES após strip (strip já foi feito antes)."""
    return MAPA_VARIACOES.get(nome, nome)


def construir_indice_colunas(header: tuple,
                             preferir_ultima: frozenset = frozenset()) -> dict:
    """
    Constrói {nome_canônico: índice_posição}.

    Pipeline por célula:
      1. strip()            — remove espaços iniciais/finais
      2. normalizar_col()   — aplica variações conhecidas
      3. None → ignorado
      4. Duplicata → primeira ocorrência vence, EXCETO para nomes em
         `preferir_ultima`, cuja ÚLTIMA ocorrência vence (D2: em 2025 a
         posição ~38 é a correta para "Classificação assistencial").
    """
    idx = {}
    for i, raw in enumerate(header):
        nome = strip_col(raw)
        if nome is None:
            continue
        nome = normalizar_col(nome)
        if nome not in idx or nome in preferir_ultima:
            idx[nome] = i
    return idx


def selecionar_aba_estado(wb) -> str:
    """
    Retorna a aba que NÃO contém 'capital' no nome (case-insensitive).
    Falha alto se não houver exatamente uma aba não-Capital.
    """
    nao_capital = [s for s in wb.sheetnames if "capital" not in s.lower()]
    assert len(nao_capital) == 1, (
        f"Esperava 1 aba não-Capital; encontrou {len(nao_capital)}: {nao_capital}"
    )
    return nao_capital[0]


def validar_colunas(col_idx: dict, ano: int, arquivo: str):
    """
    Portão de validação: todas as 33 colunas canônicas devem estar presentes.
    Imprime diagnóstico detalhado e lança AssertionError se faltar alguma.
    """
    ausentes = [c for c in COLUNAS_CANONICAS if c not in col_idx]
    if ausentes:
        print(f"\n[ERRO] {arquivo} (ano={ano}) — colunas canônicas ausentes:")
        for c in ausentes:
            print(f"       • '{c}'")
    assert not ausentes, (
        f"{arquivo}: {len(ausentes)} colunas ausentes. Interrompa para inspeção manual."
    )


def val_float(x) -> float:
    """Converte para float; retorna 0.0 se None ou não conversível."""
    if x is None:
        return 0.0
    try:
        return float(x)
    except (ValueError, TypeError):
        return 0.0


def eh_obito(desfecho) -> bool:
    """True se DESFECHO começa com 'Óbito' (NFC-normalizado)."""
    if not isinstance(desfecho, str):
        return False
    return unicodedata.normalize("NFC", desfecho).startswith(OBITO_NFC)


def eh_obito_versao_b(desfecho) -> bool:
    """
    True se o desfecho entra na contagem de mort_sem_excl (versão B).

    Usa DESFECHO_OBITO_VERSAO_B para decisão explícita por categoria.
    Se o valor não estiver no dicionário (categoria nova/desconhecida):
      - loga aviso uma única vez
      - inclui por default (conservador: não perde óbito silenciosamente)
    """
    if not isinstance(desfecho, str):
        return False
    d = unicodedata.normalize("NFC", desfecho)
    if d in DESFECHO_OBITO_VERSAO_B:
        return DESFECHO_OBITO_VERSAO_B[d]
    if d.startswith(OBITO_NFC):
        print(f"[AVISO] DESFECHO Óbito desconhecido — incluído na versão B por default: '{d}'")
    return d.startswith(OBITO_NFC)


def eh_covid(cod_grupo, cod_sub) -> bool:
    """
    True se o procedimento está marcado como COVID
    (código 999 em Cód Grupo ou Cód Subgrupo). Aplicável em 2020 e 2021.
    """
    return cod_grupo == 999 or cod_sub == 999


# ══════════════════════════════════════════════════════════════════════════════
# D. STREAMING E AGREGAÇÃO POR ARQUIVO
# ══════════════════════════════════════════════════════════════════════════════

def _prod0() -> dict:
    """Acumulador vazio para linhas de produção de um CNES."""
    d = {
        "qtde": 0.0, "qtde_obito_all": 0.0, "qtde_obito_sem_excl": 0.0,
        "dias_perm": 0.0, "valor": 0.0, "qtde_alta_complex": 0.0,
        "qtde_covid": 0.0, "dias_covid": 0.0, "valor_covid": 0.0,
        "n_linhas": 0, "n_linhas_covid": 0,
    }
    if ITEM_113_MORT_ESTRATIFICADA:
        d["qtde_obito_alta_complex"] = 0.0
    return d


def _resumo0() -> dict:
    """Acumulador vazio para a linha-resumo de um CNES."""
    return {
        "nome_fantasia": None, "class_assistencial": None,
        "porte_hospital": None, "total_leitos": None, "total_leitos_sus": None,
        "leitos_internacao": None, "leitos_sus": None,
        "uti_total": None, "uti_sus": None,
        "diarias_internacao": None, "diarias_uti": None,
        "ocupacao_internacao": None, "ocupacao_uti": None,
        "cod_ibge": None, "municipio": None,
        "tipo_hospital": None, "especializacao": None,
        "n_linhas_resumo": 0,
    }


def processar_arquivo_sih(path: Path, ano: int) -> tuple:
    """
    Lê um arquivo SIH em modo streaming (openpyxl read_only + data_only).
    Nunca carrega todos os dados em memória: agrega por CNES linha a linha.

    Regra de separação (sem linhas híbridas por premissa verificada):
        RESUMO   : QTDE é None  E  Total Leitos não é None
        PRODUÇÃO : QTDE não é None
        IGNORADA : ambos None (linha vazia ou irrelevante)

    Retorna (prod_dict, resumo_dict, log_dict).
    """
    arquivo = path.name
    print(f"\n[STREAM] {arquivo}  (ano={ano}) ...")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    aba = selecionar_aba_estado(wb)
    print(f"         Aba: '{aba}'")
    ws = wb[aba]

    linhas = ws.iter_rows(values_only=True)

    # Lê cabeçalho e constrói índice
    # D2: só em 2025 a última ocorrência de "Classificação assistencial" vence
    header = next(linhas)
    preferir = (frozenset({"Classificação assistencial"}) if ano == 2025
                else frozenset())
    col_idx = construir_indice_colunas(header, preferir_ultima=preferir)
    validar_colunas(col_idx, ano, arquivo)

    # Atalhos de índice (acesso por nome; nenhum por posição fixa)
    gi = col_idx.__getitem__
    i_cnes    = gi("CNES");                i_nome    = gi("NOME FANTASIA")
    i_qtde    = gi("QTDE");                i_dias    = gi("SomaDeDIAS_PERM")
    i_valor   = gi("VALOR TOTAL");         i_desfech = gi("DESFECHO")
    i_complex = gi("COMPLEX");             i_cgrp    = gi("Cód Grupo")
    i_csub    = gi("Cód Subgrupo");        i_class_a = gi("Classificação assistencial")
    i_leit_i  = gi("Leitos Internação");   i_leit_s  = gi("Leitos SUS")
    i_uti_t   = gi("UTI");                 i_uti_s   = gi("UTI SUS")
    i_tot_l   = gi("Total Leitos");        i_tot_ls  = gi("Total Leitos SUS")
    i_diar_i  = gi("Diárias Internação Ano"); i_diar_u = gi("Diárias UTI Ano")
    i_porte   = gi("Porte Hospital")
    i_ocup_i  = gi("Ocupação Internação"); i_ocup_u  = gi("Ocupação UTI")
    i_ibge    = gi("Cód IBGE");            i_mun     = gi("Município")
    i_tipo    = gi("Tipo de Hospital");    i_espec   = gi("Especialização")

    prod_dict   = {}
    resumo_dict = {}
    n_prod = n_resumo = n_hibrido = n_vazias = 0
    proc_covid = (ano in ANOS_COVID)

    # D2 (2025): a classificação correta é uma TABELA DE LOOKUP embutida —
    # pares (CNES na posição i_class_a-1 → rótulo na posição i_class_a) nas
    # primeiras ~637 linhas de produção. Na linha-resumo essa posição é None,
    # então o rótulo é atribuído por CNES ao final do streaming.
    usa_lookup_class = bool(preferir)
    lookup_class: dict = {}

    for row in linhas:
        if usa_lookup_class and row[i_class_a] is not None \
                and row[i_class_a - 1] is not None:
            lookup_class[row[i_class_a - 1]] = row[i_class_a]

        # Ignora linhas com CNES vazio
        cnes = row[i_cnes]
        if cnes is None:
            n_vazias += 1
            continue

        qtde      = row[i_qtde]
        tot_leitos = row[i_tot_l]

        eh_resumo = (qtde is None)     and (tot_leitos is not None)
        eh_prod   = (qtde is not None)

        # Linha híbrida: não deveria existir — reportada mas não processa
        if eh_resumo and eh_prod:
            n_hibrido += 1
            continue

        # ── RESUMO (1 linha por CNES): dados de leitos/ocupação ─────────────
        if eh_resumo:
            n_resumo += 1
            if cnes not in resumo_dict:
                resumo_dict[cnes] = _resumo0()
            r = resumo_dict[cnes]
            r["n_linhas_resumo"] += 1
            if r["n_linhas_resumo"] == 1:   # só preenche na primeira linha
                r["nome_fantasia"]       = row[i_nome]
                r["class_assistencial"]  = row[i_class_a]
                r["porte_hospital"]      = row[i_porte]
                r["total_leitos"]        = row[i_tot_l]
                r["total_leitos_sus"]    = row[i_tot_ls]
                r["leitos_internacao"]   = row[i_leit_i]
                r["leitos_sus"]          = row[i_leit_s]
                r["uti_total"]           = row[i_uti_t]
                r["uti_sus"]             = row[i_uti_s]
                r["diarias_internacao"]  = row[i_diar_i]
                r["diarias_uti"]         = row[i_diar_u]
                r["ocupacao_internacao"] = row[i_ocup_i]
                r["ocupacao_uti"]        = row[i_ocup_u]
                r["cod_ibge"]            = row[i_ibge]
                r["municipio"]           = row[i_mun]
                r["tipo_hospital"]       = row[i_tipo]
                r["especializacao"]      = row[i_espec]
            continue

        # ── PRODUÇÃO: dados de procedimentos ─────────────────────────────────
        if not eh_prod:
            # QTDE=None e Total Leitos=None — linha irrelevante
            continue

        n_prod += 1
        q = val_float(qtde)
        d = val_float(row[i_dias])
        v = val_float(row[i_valor])
        desfecho = row[i_desfech]
        complex_ = row[i_complex]
        cod_grp  = row[i_cgrp]
        cod_sub  = row[i_csub]

        if cnes not in prod_dict:
            prod_dict[cnes] = _prod0()
        p = prod_dict[cnes]

        p["qtde"]     += q
        p["dias_perm"] += d
        p["valor"]    += v
        p["n_linhas"] += 1

        if eh_obito(desfecho):
            p["qtde_obito_all"] += q
            if eh_obito_versao_b(desfecho):
                p["qtde_obito_sem_excl"] += q
            # item 1.13: óbito e complexidade estão na mesma linha — cruzamento
            if ITEM_113_MORT_ESTRATIFICADA and complex_ == "Alta complexidade":
                p["qtde_obito_alta_complex"] += q

        if complex_ == "Alta complexidade":
            p["qtde_alta_complex"] += q

        if proc_covid and eh_covid(cod_grp, cod_sub):
            p["qtde_covid"]     += q
            p["dias_covid"]     += d
            p["valor_covid"]    += v
            p["n_linhas_covid"] += 1

    wb.close()

    # D2 (2025): atribui o rótulo correto por CNES a partir do lookup
    if usa_lookup_class:
        atribuidos = 0
        for cnes_r, r in resumo_dict.items():
            rotulo = lookup_class.get(cnes_r)
            if rotulo is not None:
                r["class_assistencial"] = rotulo
                atribuidos += 1
        print(f"         D2/2025: lookup CNES→classificação com "
              f"{len(lookup_class)} pares; {atribuidos}/{len(resumo_dict)} "
              f"CNES do resumo com rótulo atribuído")

    # Valida: exatamente 1 linha-resumo por CNES
    multiplos = {c: r["n_linhas_resumo"]
                 for c, r in resumo_dict.items() if r["n_linhas_resumo"] > 1}
    if multiplos:
        print(f"  [ALERTA] {len(multiplos)} CNES com >1 linha-resumo: "
              f"{list(multiplos.items())[:5]}")

    log = {
        "ano": ano, "arquivo": arquivo,
        "n_prod": n_prod, "n_resumo": n_resumo,
        "n_hibrido": n_hibrido, "n_vazias": n_vazias,
        "n_cnes_prod": len(prod_dict), "n_cnes_resumo": len(resumo_dict),
    }
    aviso = f" | HÍBRIDAS: {n_hibrido} [!!!]" if n_hibrido else ""
    print(
        f"         Produção: {n_prod:>9,} linhas  {len(prod_dict):>4} CNES"
        f"   |   Resumo: {n_resumo:>4} linhas  {len(resumo_dict):>4} CNES"
        + aviso
    )
    return prod_dict, resumo_dict, log


# ══════════════════════════════════════════════════════════════════════════════
# E. CONSTRUÇÃO DO DATAFRAME POR ANO E DO PAINEL COMPLETO
# ══════════════════════════════════════════════════════════════════════════════

def construir_df_arquivo(prod_dict: dict, resumo_dict: dict, ano: int) -> pd.DataFrame:
    """
    Combina produção + resumo em 1 linha por CNES.
    CNES só na produção → colunas de resumo = NaN (e vice-versa).
    Pandas usado apenas aqui, sobre os ~638 CNES agregados.
    """
    todos_cnes = sorted(set(prod_dict) | set(resumo_dict))
    registros = []
    for cnes in todos_cnes:
        p = prod_dict.get(cnes,   _prod0())
        r = resumo_dict.get(cnes, _resumo0())
        registros.append({
            "cnes": cnes, "ano": ano,
            # --- do resumo ---
            "nome_fantasia":       r["nome_fantasia"],
            "class_assistencial":  r["class_assistencial"],
            "porte_hospital":      r["porte_hospital"],
            "total_leitos":        r["total_leitos"],
            "total_leitos_sus":    r["total_leitos_sus"],
            "leitos_internacao":   r["leitos_internacao"],
            "leitos_sus":          r["leitos_sus"],
            "uti_total":           r["uti_total"],
            "uti_sus":             r["uti_sus"],
            "diarias_internacao":  r["diarias_internacao"],
            "diarias_uti":         r["diarias_uti"],
            "ocupacao_internacao": r["ocupacao_internacao"],  # já calculada; não recalcular
            "ocupacao_uti":        r["ocupacao_uti"],
            "cod_ibge":            r["cod_ibge"],
            "municipio":           r["municipio"],
            "tipo_hospital":       r["tipo_hospital"],
            "especializacao":      r["especializacao"],
            "n_linhas_resumo":     r["n_linhas_resumo"],
            # --- da produção ---
            "qtde":                p["qtde"],
            "qtde_obito_all":      p["qtde_obito_all"],
            "qtde_obito_sem_excl": p["qtde_obito_sem_excl"],
            "dias_perm":           p["dias_perm"],
            "valor":               p["valor"],
            "qtde_alta_complex":   p["qtde_alta_complex"],
            **({"qtde_obito_alta_complex": p["qtde_obito_alta_complex"]}
               if ITEM_113_MORT_ESTRATIFICADA else {}),
            "qtde_covid":          p["qtde_covid"],
            "dias_covid":          p["dias_covid"],
            "valor_covid":         p["valor_covid"],
            "n_linhas_prod":       p["n_linhas"],
            "n_linhas_covid":      p["n_linhas_covid"],
        })
    return pd.DataFrame(registros)


def construir_painel_completo(arquivos_sih: list) -> tuple:
    """
    Processa todos os SIH em streaming e empilha os DataFrames anuais.
    Retorna (painel, lista_de_logs).
    """
    frames, logs = [], []
    for path, ano in arquivos_sih:
        prod, resumo, log = processar_arquivo_sih(path, ano)
        frames.append(construir_df_arquivo(prod, resumo, ano))
        logs.append(log)
    painel = pd.concat(frames, ignore_index=True)
    print(f"\n[PAINEL] Shape: {painel.shape}  "
          f"({painel['cnes'].nunique()} CNES únicos | "
          f"{painel['ano'].nunique()} anos)")
    return painel, logs


def salvar_painel(painel: pd.DataFrame):
    """Salva painel em parquet (preferencial) ou CSV (fallback sem pyarrow)."""
    try:
        painel.to_parquet(PAINEL_PARQUET, index=False)
        print(f"[SAÍDA] Painel → {PAINEL_PARQUET}")
    except Exception:
        painel.to_csv(PAINEL_CSV, index=False, encoding="utf-8-sig")
        print(f"[SAÍDA] pyarrow indisponível; painel → {PAINEL_CSV}")


def carregar_painel_cache():
    """Carrega painel de cache para evitar reprocessamento; retorna None se não houver."""
    if PAINEL_PARQUET.exists():
        df = pd.read_parquet(PAINEL_PARQUET)
        print(f"[CACHE] Painel carregado de {PAINEL_PARQUET.name}  shape={df.shape}")
        return df
    if PAINEL_CSV.exists():
        df = pd.read_csv(PAINEL_CSV, encoding="utf-8-sig")
        print(f"[CACHE] Painel carregado de {PAINEL_CSV.name}  shape={df.shape}")
        return df
    return None


# ══════════════════════════════════════════════════════════════════════════════
# F. PLANILHA DE CLASSIFICAÇÃO HOSPITALAR
# ══════════════════════════════════════════════════════════════════════════════

def carregar_classificacao(path: Path) -> pd.DataFrame:
    """
    Carrega a planilha de classificação hospitalar (aba "Com Municípios").
    Strip em todos os cabeçalhos resolve "UF " → "UF" (D3).
    """
    ABA = "Com Municípios"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    assert ABA in wb.sheetnames, (
        f"Aba '{ABA}' não encontrada em {path.name}. Abas: {wb.sheetnames}"
    )
    ws = wb[ABA]
    linhas = list(ws.iter_rows(values_only=True))
    wb.close()
    assert linhas, f"Aba '{ABA}' vazia."

    # Strip universal nos cabeçalhos (resolve D3: "UF " → "UF")
    header = [strip_col(c) for c in linhas[0]]

    ausentes = [c for c in COLUNAS_CLASSIF if c not in set(header)]
    if ausentes:
        print(f"[ERRO] Classificação: colunas ausentes após strip: {ausentes}")
    assert not ausentes, "Planilha de classificação: colunas faltando."

    df = pd.DataFrame(linhas[1:], columns=header).dropna(how="all").reset_index(drop=True)
    print(f"[CLASSIF] {path.name}: {len(df)} hospitais carregados.")
    return df


def validar_formula_barcelona(df: pd.DataFrame):
    """
    Reproduz a Pontuação Barcelona e verifica contra o valor gravado.

    Fórmula: Leitos*1 + Salas*2 + UTI*4 + 25*(Torácica>0) + 25*(Neuro>0) + 25*(Pediatrica>0)
    As colunas de cirurgia são volumes, mas entram como FLAG binária (>0), não como volume.
    """
    df = df.copy()
    num_cols = ["Leitos", "Salas Cirúrgicas", "UTI",
                "Cirurgia Torácica", "Neurocirurgia", "Cirurgia Pediátrica", "Pontuação"]
    for c in num_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    df["pont_calc"] = (
          df["Leitos"].fillna(0)               * 1
        + df["Salas Cirúrgicas"].fillna(0)     * 2
        + df["UTI"].fillna(0)                  * 4
        + 25 * (df["Cirurgia Torácica"].fillna(0)  > 0).astype(int)
        + 25 * (df["Neurocirurgia"].fillna(0)      > 0).astype(int)
        + 25 * (df["Cirurgia Pediátrica"].fillna(0) > 0).astype(int)
    )

    batem = (df["pont_calc"] == df["Pontuação"]).sum()
    n     = len(df)
    print(f"[BARCELONA] Fórmula reproduzida: {batem}/{n} hospitais.")
    if batem < n:
        div = df[df["pont_calc"] != df["Pontuação"]][
            ["CNES", "Pontuação", "pont_calc"]].head(10)
        print(f"  Primeiros divergentes:\n{div.to_string(index=False)}")
    assert batem > 0, "Fórmula Barcelona não reproduziu nenhum hospital."


# ══════════════════════════════════════════════════════════════════════════════
# G. DIAGNÓSTICO DE COBERTURA DO PAINEL
# ══════════════════════════════════════════════════════════════════════════════

def diagnostico_cobertura(painel: pd.DataFrame,
                          df_classif: pd.DataFrame) -> pd.DataFrame:
    """
    (1) Nº de CNES por ano.
    (2) Cruzamento SIH × Classificação: casados / não-casados por sentido e ano.
    (3) Padrão de presença CNES × ano (sempre / parcial).

    Gera fig01 e fig02; retorna DataFrame de cobertura por ano.
    """
    anos = sorted(painel["ano"].unique())
    cnes_classif = set(df_classif["CNES"].dropna().astype(int))

    resumo = []
    for ano in anos:
        cnes_ano = set(painel.loc[painel["ano"] == ano, "cnes"].astype(int))
        casados     = len(cnes_ano & cnes_classif)
        so_sih      = len(cnes_ano - cnes_classif)
        so_classif  = len(cnes_classif - cnes_ano)
        resumo.append({
            "ano":             ano,
            "n_cnes_sih":      len(cnes_ano),
            "n_casados":       casados,
            "n_so_sih":        so_sih,
            "n_so_classif":    so_classif,
            "pct_cobert_sih":  round(100 * casados / len(cnes_ano), 1) if cnes_ano else 0,
        })

    df_cob = pd.DataFrame(resumo)
    print("\n[COBERTURA POR ANO]")
    print(df_cob.to_string(index=False))

    # Padrão de presença
    presenca = painel.pivot_table(
        index="cnes", columns="ano", values="qtde",
        aggfunc=lambda x: 1, fill_value=0
    )
    n_anos = len(anos)
    n_sempre = (presenca.sum(axis=1) == n_anos).sum()
    n_parcial = (presenca.sum(axis=1) < n_anos).sum()
    print(f"\n  CNES presentes em todos os {n_anos} anos : {n_sempre}")
    print(f"  CNES com presença parcial (entrada/saída): {n_parcial}")

    # ── fig01: barras empilhadas por ano ──────────────────────────────────
    fig, axes = plt.subplots(1, 2, figsize=(13, 5))
    x = [str(a) for a in df_cob["ano"]]

    ax = axes[0]
    ax.bar(x, df_cob["n_cnes_sih"],  color="#4e79a7", label="Total SIH")
    ax.bar(x, df_cob["n_casados"],   color="#f28e2b", alpha=0.75,
           label="Com classificação Barcelona")
    ax.set_title("Nº de hospitais por ano"); ax.set_xlabel("Ano")
    ax.set_ylabel("Nº de CNES"); ax.legend(fontsize=8)
    ax.tick_params(axis="x", rotation=45)

    ax = axes[1]
    ax.plot(x, df_cob["n_so_sih"],     "o-", color="#e15759",
            label="Só no SIH (sem classif.)")
    ax.plot(x, df_cob["n_so_classif"], "s--", color="#76b7b2",
            label="Só na classif. (sem prod.)")
    ax.set_title("Não-casados por sentido"); ax.set_xlabel("Ano")
    ax.set_ylabel("Nº de CNES"); ax.legend(fontsize=8)
    ax.tick_params(axis="x", rotation=45)

    fig.suptitle("Diagnóstico de cobertura do painel")
    fig.tight_layout()
    _salvar_fig(fig, "fig01_cobertura_painel.png")

    # ── fig02: distribuição de CNES por nº de anos presentes ─────────────
    contagem = presenca.sum(axis=1).value_counts().sort_index()
    fig2, ax2 = plt.subplots(figsize=(7, 4))
    ax2.bar([str(k) for k in contagem.index], contagem.values, color="#59a14f")
    ax2.set_title("Distribuição de CNES por nº de anos com produção")
    ax2.set_xlabel("Anos presentes"); ax2.set_ylabel("Nº de CNES")
    fig2.tight_layout()
    _salvar_fig(fig2, "fig02_anos_presentes_por_cnes.png")

    return df_cob


# ══════════════════════════════════════════════════════════════════════════════
# H. INDICADORES HOSPITAL-ANO
# ══════════════════════════════════════════════════════════════════════════════

def _div(num, den):
    """Divisão vetorizada segura; retorna NaN onde denominador <= 0."""
    return np.where(den > 0, num / den, np.nan)


def calcular_indicadores(painel: pd.DataFrame) -> pd.DataFrame:
    """
    Adiciona colunas de indicadores ao painel.

    mort_all        : todos os óbitos / qtde total
    mort_sem_excl   : exclui óbito fetal/materno/neonatal (versão B — decisão clínica)
    tmp             : dias_perm / qtde  (tempo médio de permanência)
    custo_saida     : valor / qtde      (faturamento por saída, R$)
    pct_alta_complex: qtde_alta_complex / qtde
    ocupacao_*      : vem do resumo — NÃO recalculada aqui
    pct_covid_*     : só preenchida em 2020-2021; NaN nos demais anos
    """
    df = painel.copy()

    df["mort_all"]        = _div(df["qtde_obito_all"],      df["qtde"])
    df["mort_sem_excl"]   = _div(df["qtde_obito_sem_excl"], df["qtde"])
    df["tmp"]             = _div(df["dias_perm"],           df["qtde"])
    df["custo_saida"]     = _div(df["valor"],               df["qtde"])
    df["pct_alta_complex"]= _div(df["qtde_alta_complex"],   df["qtde"])

    # COVID: preenchido apenas nos anos relevantes
    mask = df["ano"].isin(ANOS_COVID)
    df["pct_covid_qtde"]  = np.nan
    df["pct_covid_valor"] = np.nan
    if mask.any():
        df.loc[mask, "pct_covid_qtde"]  = _div(
            df.loc[mask, "qtde_covid"],  df.loc[mask, "qtde"])
        df.loc[mask, "pct_covid_valor"] = _div(
            df.loc[mask, "valor_covid"], df.loc[mask, "valor"])

    return df


# ══════════════════════════════════════════════════════════════════════════════
# I. ANÁLISE COVID
# ══════════════════════════════════════════════════════════════════════════════

def relatorio_covid(painel_ind: pd.DataFrame):
    """
    Para 2020 e 2021, quantifica o peso dos procedimentos COVID (código 999)
    e compara indicadores COM e SEM esses procedimentos.
    """
    anos_presentes = [a for a in sorted(ANOS_COVID)
                      if a in painel_ind["ano"].unique()]
    if not anos_presentes:
        print("[COVID] Nenhum ano COVID no painel.")
        return

    # ── Peso do COVID no total ────────────────────────────────────────────
    linhas = []
    for ano in anos_presentes:
        sub = painel_ind[painel_ind["ano"] == ano]
        qtde_tot  = sub["qtde"].sum()
        qtde_cov  = sub["qtde_covid"].sum()
        valor_tot = sub["valor"].sum()
        valor_cov = sub["valor_covid"].sum()
        linhas.append({
            "ano":                ano,
            "qtde_total":         qtde_tot,
            "qtde_covid":         qtde_cov,
            "pct_qtde_covid":     100 * qtde_cov  / qtde_tot  if qtde_tot  else 0,
            "valor_total_R$":     valor_tot,
            "valor_covid_R$":     valor_cov,
            "pct_valor_covid":    100 * valor_cov / valor_tot if valor_tot else 0,
            "n_hosp_com_covid":   (sub["qtde_covid"] > 0).sum(),
        })
    df_peso = pd.DataFrame(linhas)
    print("\n[COVID] Peso dos procedimentos COVID (código 999):")
    print(df_peso.to_string(index=False))
    df_peso.to_csv(PASTA_TABELAS / "tab_covid_peso.csv",
                   index=False, encoding="utf-8-sig")

    # ── Comparação indicadores COM vs SEM COVID ───────────────────────────
    comp = []
    for ano in anos_presentes:
        sub = painel_ind[(painel_ind["ano"] == ano) & (painel_ind["qtde"] > 0)].copy()

        # subtrai numeradores COVID para obter versão "sem COVID"
        sub["qtde_s"]  = sub["qtde"]     - sub["qtde_covid"]
        sub["dias_s"]  = sub["dias_perm"]- sub["dias_covid"]
        sub["valor_s"] = sub["valor"]    - sub["valor_covid"]
        sub["tmp_s"]   = _div(sub["dias_s"],  sub["qtde_s"])
        sub["cst_s"]   = _div(sub["valor_s"], sub["qtde_s"])

        for versao, tmp_col, cst_col, mort_col in [
            ("Com COVID",  "tmp",   "custo_saida", "mort_all"),
            ("Sem COVID",  "tmp_s", "cst_s",       "mort_all"),
        ]:
            comp.append({
                "ano": ano, "versao": versao,
                "tmp_mediana":   sub[tmp_col].median(),
                "custo_mediana": sub[cst_col].median(),
                "mort_mediana":  sub[mort_col].median(),
            })
    df_comp = pd.DataFrame(comp)
    df_comp.to_csv(PASTA_TABELAS / "tab_covid_com_sem.csv",
                   index=False, encoding="utf-8-sig")

    # ── fig03: comparação visual ──────────────────────────────────────────
    metricas = [("tmp_mediana",   "TMP mediano (dias)"),
                ("custo_mediana", "Faturamento/saída mediano (R$)"),
                ("mort_mediana",  "Mortalidade mediana")]
    fig, axes = plt.subplots(1, 3, figsize=(14, 5))
    cores = {"Com COVID": "#e15759", "Sem COVID": "#4e79a7"}
    for ax, (col, titulo) in zip(axes, metricas):
        for versao, cor in cores.items():
            s = df_comp[df_comp["versao"] == versao]
            ax.bar([f"{a}\n{versao}" for a in s["ano"]],
                   s[col], color=cor, alpha=0.85, label=versao)
        ax.set_title(titulo, fontsize=9); ax.legend(fontsize=7)
    fig.suptitle("Impacto COVID: indicadores com e sem código 999 (2020–2021)")
    fig.tight_layout()
    _salvar_fig(fig, "fig03_covid_com_sem.png")


# ══════════════════════════════════════════════════════════════════════════════
# J. ESTATÍSTICA DESCRITIVA E FIGURAS
# ══════════════════════════════════════════════════════════════════════════════

def _salvar_fig(fig, nome: str):
    """Salva figura em PASTA_FIGURAS e fecha."""
    caminho = PASTA_FIGURAS / nome
    fig.savefig(caminho, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  [FIG] {nome}")


def _melt_ind(painel_ind: pd.DataFrame) -> pd.DataFrame:
    """Retorna painel em formato longo para os 5 indicadores principais."""
    cols_ok = ["cnes", "ano"] + [c for c in INDICADORES if c in painel_ind.columns]
    return painel_ind[cols_ok].melt(
        id_vars=["cnes", "ano"],
        value_vars=[c for c in INDICADORES if c in painel_ind.columns],
        var_name="indicador", value_name="valor"
    )


def distribuicoes_indicadores(painel_ind: pd.DataFrame):
    """
    fig04: boxplots dos 5 indicadores por ano.
    fig05: histogramas gerais (todos os anos somados).
    """
    anos_ord = sorted(painel_ind["ano"].unique())
    n = len(INDICADORES)

    # fig04
    fig, axes = plt.subplots(n, 1, figsize=(14, 4 * n))
    for ax, ind in zip(axes, INDICADORES):
        sub = painel_ind[["ano", ind]].dropna()
        sns.boxplot(data=sub, x="ano", y=ind, ax=ax,
                    order=anos_ord, palette="Blues_d",
                    fliersize=2, linewidth=0.8)
        ax.set_title(ROTULOS[ind], fontsize=9)
        ax.set_xlabel("Ano"); ax.set_ylabel("")
    fig.suptitle("Distribuição dos indicadores por ano", y=1.01)
    fig.tight_layout()
    _salvar_fig(fig, "fig04_boxplots_por_ano.png")

    # fig05
    fig2, axes2 = plt.subplots(1, n, figsize=(5 * n, 4))
    for ax, ind in zip(axes2, INDICADORES):
        sub = painel_ind[ind].dropna()
        ax.hist(sub, bins=50, color="#4e79a7", edgecolor="white", linewidth=0.3)
        ax.set_title(ROTULOS[ind], fontsize=8)
        ax.set_ylabel("Nº hospitais-ano")
    fig2.suptitle("Histogramas gerais dos indicadores (todos os anos)")
    fig2.tight_layout()
    _salvar_fig(fig2, "fig05_histogramas_gerais.png")


def series_temporais(painel_ind: pd.DataFrame):
    """fig06: séries temporais das medianas ± IQR por ano."""
    grp  = painel_ind.groupby("ano")[INDICADORES]
    med  = grp.median()
    q25  = grp.quantile(0.25)
    q75  = grp.quantile(0.75)
    n    = len(INDICADORES)

    fig, axes = plt.subplots(1, n, figsize=(5 * n, 4))
    for ax, ind in zip(axes, INDICADORES):
        x = med.index
        ax.plot(x, med[ind], "o-", color="#4e79a7", lw=1.8, ms=5)
        ax.fill_between(x, q25[ind], q75[ind], alpha=0.2, color="#4e79a7")
        ax.set_title(ROTULOS[ind], fontsize=8)
        ax.set_xlabel("Ano"); ax.tick_params(axis="x", rotation=45)
    fig.suptitle("Séries temporais — mediana ± IQR")
    fig.tight_layout()
    _salvar_fig(fig, "fig06_series_temporais.png")


def cortes_complexidade_porte(painel_ind: pd.DataFrame,
                               df_classif: pd.DataFrame):
    """
    fig07: indicadores por Faixa de complexidade Barcelona.
    fig08: indicadores por Porte Hospital (SIH).

    Nota DRS: mapeamento município→DRS não disponível nos dados.
    Para incluir esse corte, forneça um CSV com [cod_ibge, drs].
    """
    # Adiciona Faixa de complexidade Barcelona ao painel
    cl = df_classif[["CNES", "Classificação"]].copy()
    cl.columns = ["cnes", "faixa_complex"]
    try:
        cl["cnes"] = cl["cnes"].astype(painel_ind["cnes"].dtype)
    except Exception:
        pass
    df_m = painel_ind.merge(cl, on="cnes", how="left")

    for col_corte, titulo, nome_fig in [
        ("faixa_complex",   "Faixa de complexidade Barcelona (2–6)", "fig07_por_complexidade"),
        ("porte_hospital",  "Porte Hospital (SIH)",                  "fig08_por_porte"),
    ]:
        cats = sorted(df_m[col_corte].dropna().unique())
        if not cats:
            continue
        n = len(INDICADORES)
        fig, axes = plt.subplots(1, n, figsize=(5 * n, 5))
        for ax, ind in zip(axes, INDICADORES):
            vals = [df_m.loc[df_m[col_corte] == c, ind].dropna() for c in cats]
            ax.boxplot(vals, labels=[str(c) for c in cats],
                       patch_artist=True, flierprops={"markersize": 2})
            ax.set_title(ROTULOS[ind], fontsize=8)
            ax.tick_params(axis="x", rotation=45, labelsize=7)
        fig.suptitle(f"Indicadores por {titulo}")
        fig.tight_layout()
        _salvar_fig(fig, f"{nome_fig}.png")


def fronteiras_proporcoes(painel_ind: pd.DataFrame):
    """
    fig09: comportamento nas fronteiras das proporções.
    Mortalidade próxima de 0; ocupação próxima/acima de 100%.
    Orienta futura escolha de modelo Beta/ZOIB.
    """
    fig, axes = plt.subplots(1, 3, figsize=(15, 4))

    # Mortalidade — zoom em [0, 15%]
    ax = axes[0]
    mort = painel_ind["mort_all"].dropna()
    ax.hist(mort[mort <= 0.15], bins=60,
            color="#e15759", edgecolor="white", lw=0.3)
    pct0 = 100 * (mort == 0).sum() / len(mort) if len(mort) else 0
    ax.set_title("Mortalidade (zoom 0–15%)")
    ax.set_xlabel("mort_all"); ax.set_ylabel("Nº hospitais-ano")
    ax.text(0.97, 0.95, f"{pct0:.1f}% com mort=0",
            transform=ax.transAxes, ha="right", va="top", fontsize=8)

    # Ocupação internação
    ax = axes[1]
    ocup_i = painel_ind["ocupacao_internacao"].dropna()
    ax.hist(ocup_i, bins=60, color="#59a14f", edgecolor="white", lw=0.3)
    ax.axvline(1.0, color="red", lw=1.5, ls="--", label=">100%")
    if len(ocup_i):
        pct = 100 * (ocup_i > 1).sum() / len(ocup_i)
        ax.text(0.97, 0.95, f"{pct:.1f}% > 100%",
                transform=ax.transAxes, ha="right", va="top", fontsize=8)
    ax.set_title("Ocupação internação"); ax.set_xlabel("Ocupação")
    ax.legend(fontsize=8)

    # Ocupação UTI
    ax = axes[2]
    ocup_u = painel_ind["ocupacao_uti"].dropna()
    ax.hist(ocup_u, bins=60, color="#76b7b2", edgecolor="white", lw=0.3)
    ax.axvline(1.0, color="red", lw=1.5, ls="--", label=">100%")
    if len(ocup_u):
        pct = 100 * (ocup_u > 1).sum() / len(ocup_u)
        ax.text(0.97, 0.95, f"{pct:.1f}% > 100%",
                transform=ax.transAxes, ha="right", va="top", fontsize=8)
    ax.set_title("Ocupação UTI"); ax.set_xlabel("Ocupação")
    ax.legend(fontsize=8)

    fig.suptitle("Comportamento nas fronteiras das proporções")
    fig.tight_layout()
    _salvar_fig(fig, "fig09_fronteiras_proporcoes.png")


def matriz_correlacao(painel_ind: pd.DataFrame):
    """fig10: correlação de Spearman entre indicadores principais."""
    cols = INDICADORES + ["ocupacao_internacao", "ocupacao_uti"]
    cols = [c for c in cols if c in painel_ind.columns]
    corr = painel_ind[cols].corr(method="spearman")
    corr.index   = [ROTULOS.get(c, c) for c in corr.index]
    corr.columns = [ROTULOS.get(c, c) for c in corr.columns]

    fig, ax = plt.subplots(figsize=(9, 7))
    sns.heatmap(corr, annot=True, fmt=".2f", cmap="RdBu_r",
                center=0, ax=ax, linewidths=0.5, vmin=-1, vmax=1)
    ax.set_title("Correlação de Spearman entre indicadores (todos os anos)")
    fig.tight_layout()
    _salvar_fig(fig, "fig10_correlacao.png")


def hospitais_extremos(painel_ind: pd.DataFrame, df_classif: pd.DataFrame):
    """
    Para cada indicador, tabula os 10 hospitais-ano nos extremos.
    Salva CSVs em PASTA_TABELAS.
    """
    cl = df_classif[["CNES", "Classificação"]].copy()
    cl.columns = ["cnes", "faixa_complex"]
    try:
        cl["cnes"] = cl["cnes"].astype(painel_ind["cnes"].dtype)
    except Exception:
        pass
    df_m = painel_ind.merge(cl, on="cnes", how="left")
    cols_id = ["cnes", "ano", "nome_fantasia", "municipio",
               "porte_hospital", "faixa_complex"]

    for ind in INDICADORES:
        sub = df_m[cols_id + [ind]].dropna(subset=[ind])
        sub.nlargest(10, ind).to_csv(
            PASTA_TABELAS / f"extremos_{ind}_top10.csv",
            index=False, encoding="utf-8-sig")
        sub.nsmallest(10, ind).to_csv(
            PASTA_TABELAS / f"extremos_{ind}_bottom10.csv",
            index=False, encoding="utf-8-sig")
    print("  [TAB] Tabelas de extremos salvas.")


def corte_municipio(painel_ind: pd.DataFrame, top_n: int = 20):
    """
    fig11: medianas dos indicadores nos top-N municípios por volume de produção.

    NOTA DRS: mapeamento município → DRS não disponível nesta base.
    Para habilitar esse corte, forneça um arquivo CSV com colunas [cod_ibge, drs]
    e adicione o merge aqui.
    """
    prod_mun = (painel_ind.groupby("municipio")["qtde"]
                .sum().nlargest(top_n).index.tolist())
    df_top = painel_ind[painel_ind["municipio"].isin(prod_mun)]
    n = len(INDICADORES)

    fig, axes = plt.subplots(n, 1, figsize=(14, 4 * n))
    for ax, ind in zip(axes, INDICADORES):
        med = (df_top.groupby("municipio")[ind].median()
               .reindex(prod_mun).reset_index())
        med.columns = ["municipio", "valor"]
        ax.barh(med["municipio"], med["valor"], color="#4e79a7")
        ax.set_title(ROTULOS[ind], fontsize=9); ax.set_xlabel("Mediana")
    fig.suptitle(f"Indicadores — top {top_n} municípios (por volume de produção)")
    fig.tight_layout()
    _salvar_fig(fig, "fig11_por_municipio.png")
    print("  [NOTA] Corte por DRS omitido: forneça CSV [cod_ibge, drs] para incluir.")


def corte_class_assistencial(painel_ind: pd.DataFrame):
    """
    fig12: indicadores por natureza jurídico-administrativa (SIH).

    Usa a coluna 'class_assistencial' (= "Classificação assistencial" do SIH).
    ATENÇÃO: a coluna mistura categorias de esferas distintas —
      • propriedade: Público Municipal, Privado
      • modelo de gestão: OSS, Direta
      • natureza jurídica: Filantrópico
    NÃO é equivalente ao "modelo de gestão" do protocolo de pesquisa.
    """
    if "class_assistencial" not in painel_ind.columns:
        return
    cats = sorted(painel_ind["class_assistencial"].dropna().unique())
    if not cats:
        return
    n = len(INDICADORES)
    fig, axes = plt.subplots(1, n, figsize=(5 * n, 6))
    for ax, ind in zip(axes, INDICADORES):
        vals = [painel_ind.loc[painel_ind["class_assistencial"] == c, ind].dropna()
                for c in cats]
        ax.boxplot(vals, labels=[str(c) for c in cats],
                   patch_artist=True, flierprops={"markersize": 2})
        ax.set_title(ROTULOS[ind], fontsize=8)
        ax.tick_params(axis="x", rotation=55, labelsize=7)
    fig.suptitle(
        "Indicadores por natureza jurídico-administrativa (SIH)\n"
        "coluna 'Classificação assistencial' — categorias mistas: "
        "propriedade / gestão / natureza jurídica — NÃO é modelo de gestão da pesquisa"
    )
    fig.tight_layout()
    _salvar_fig(fig, "fig12_por_class_assistencial.png")


# ══════════════════════════════════════════════════════════════════════════════
# K. SAÍDAS TABULARES
# ══════════════════════════════════════════════════════════════════════════════

def salvar_tabelas(painel_ind: pd.DataFrame, df_classif: pd.DataFrame,
                   df_cob: pd.DataFrame, logs: list):
    """Salva tabelas-resumo em PASTA_TABELAS."""
    # Cobertura por ano
    df_cob.to_csv(PASTA_TABELAS / "tab_cobertura_por_ano.csv",
                  index=False, encoding="utf-8-sig")

    # Descritiva dos indicadores por ano
    cols_desc = INDICADORES + ["ocupacao_internacao", "ocupacao_uti"]
    cols_desc = [c for c in cols_desc if c in painel_ind.columns]
    (painel_ind.groupby("ano")[cols_desc]
     .describe(percentiles=[.25, .5, .75])
     .round(4)
     .to_csv(PASTA_TABELAS / "tab_descritiva_por_ano.csv", encoding="utf-8-sig"))

    # Painel completo enriquecido com Barcelona
    cl = df_classif[["CNES", "Classificação", "Pontuação"]].copy()
    cl.columns = ["cnes", "faixa_complex_barcelona", "pont_barcelona"]
    try:
        cl["cnes"] = cl["cnes"].astype(painel_ind["cnes"].dtype)
    except Exception:
        pass
    painel_full = painel_ind.merge(cl, on="cnes", how="left")
    painel_full.to_csv(PASTA_TABELAS / "tab_painel_completo.csv",
                       index=False, encoding="utf-8-sig")

    # Classificação hospitais
    df_classif.to_csv(PASTA_TABELAS / "tab_classificacao_hospitais.csv",
                      index=False, encoding="utf-8-sig")

    # Log de processamento
    if logs:
        pd.DataFrame(logs).to_csv(PASTA_TABELAS / "tab_log_processamento.csv",
                                   index=False, encoding="utf-8-sig")
    print(f"[TAB] Tabelas salvas em {PASTA_TABELAS}")


# ══════════════════════════════════════════════════════════════════════════════
# L. MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 65)
    print("ANÁLISE EXPLORATÓRIA SIH/SUS — SP 2015-2025")
    print("=" * 65)

    configurar_diretorios()

    # Localizar arquivos
    arquivos_sih, path_classif, _ = localizar_arquivos(PASTA_DADOS)
    print(f"\n[ARQUIVOS] {len(arquivos_sih)} SIH: {[a for _, a in arquivos_sih]}")
    assert arquivos_sih,      "Nenhum arquivo SIH encontrado."
    assert path_classif,      "Planilha de classificação não encontrada."

    # Classificação hospitais
    df_classif = carregar_classificacao(path_classif)
    validar_formula_barcelona(df_classif)

    # Painel hospital-ano (cache ou streaming)
    logs   = []
    painel = carregar_painel_cache()
    if painel is None:
        painel, logs = construir_painel_completo(arquivos_sih)
        salvar_painel(painel)
    else:
        print("[INFO] Cache encontrado. Para reprocessar: apague o parquet/csv do painel.")

    # Cobertura
    df_cob = diagnostico_cobertura(painel, df_classif)

    # Indicadores
    painel_ind = calcular_indicadores(painel)

    # COVID
    relatorio_covid(painel_ind)

    # Descritiva e figuras
    print("\n[FIGURAS] Gerando...")
    distribuicoes_indicadores(painel_ind)
    series_temporais(painel_ind)
    cortes_complexidade_porte(painel_ind, df_classif)
    fronteiras_proporcoes(painel_ind)
    matriz_correlacao(painel_ind)
    hospitais_extremos(painel_ind, df_classif)
    corte_municipio(painel_ind, top_n=20)
    corte_class_assistencial(painel_ind)

    # Saídas tabulares
    salvar_tabelas(painel_ind, df_classif, df_cob, logs)

    print("\n" + "=" * 65)
    print("CONCLUÍDO. Verifique ./analises/")
    print("=" * 65)


if __name__ == "__main__":
    main()
