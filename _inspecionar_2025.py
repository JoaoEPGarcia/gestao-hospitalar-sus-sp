# -*- coding: utf-8 -*-
"""
Inspeção focada no arquivo 2025: verifica posição 17 (Classificação assistencial)
e a posição 37 (que deveria ser None/padding) nas primeiras linhas de dados.
"""
import io, sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
from pathlib import Path

PASTA = Path(__file__).parent
PATH_2025 = PASTA / "20260605 - 2025.xlsx"

wb = openpyxl.load_workbook(PATH_2025, read_only=True, data_only=True)
ws = wb["Estado - 2025"]
rows_iter = ws.iter_rows(values_only=True)

# ── Cabeçalho ────────────────────────────────────────────────────────────────
header = next(rows_iter)
print(f"Total de colunas no cabeçalho: {len(header)}")
print()

# Mostra cada coluna com sua posição
for i, nome in enumerate(header):
    if nome is not None:
        print(f"  idx {i:>2}: {repr(nome)}")

print()
print("=" * 60)
print("Valores de POSIÇÃO 17 e POSIÇÃO 37 nas primeiras 60 linhas de dados:")
print(f"  {'Linha':>5}  {'pos17 (Classif. assist.)':30}  {'pos37 (deveria ser None)':25}  {'pos38 (Classif. Assistencial)':30}")
print("-" * 100)

n_linhas_pos17_none = 0
n_linhas_pos37_nao_none = 0
n_total = 0

for i, row in enumerate(rows_iter):
    if i >= 60:
        break
    n_total += 1
    val17 = row[17] if len(row) > 17 else "FORA DO RANGE"
    val37 = row[37] if len(row) > 37 else "FORA DO RANGE"
    val38 = row[38] if len(row) > 38 else "FORA DO RANGE"

    if val17 is None:
        n_linhas_pos17_none += 1
    if val37 is not None:
        n_linhas_pos37_nao_none += 1

    # Exibe resumo (apenas QTDE para saber se é produção ou resumo)
    qtde = row[4] if len(row) > 4 else None
    tipo = "RESUMO" if qtde is None else "PROD  "
    print(f"  {i:>5} [{tipo}]  {repr(val17):35}  {repr(val37):25}  {repr(val38)}")

print()
print(f"Resumo das 60 primeiras linhas:")
print(f"  pos17 com None : {n_linhas_pos17_none} / {n_total}")
print(f"  pos37 não-None : {n_linhas_pos37_nao_none} / {n_total}")

# Agora varredura mais ampla: percorre TODAS as linhas e conta Nones na pos17
print()
print("Varrendo TODO o arquivo para contar Nones na pos17...")
wb.close()

wb2 = openpyxl.load_workbook(PATH_2025, read_only=True, data_only=True)
ws2 = wb2["Estado - 2025"]
rows2 = ws2.iter_rows(values_only=True)
next(rows2)  # pula cabeçalho

tot, prod_n, resumo_n = 0, 0, 0
pos17_none_prod = 0
pos17_none_resumo = 0
pos37_nao_none = 0

for row in rows2:
    cnes = row[0]
    if cnes is None:
        continue
    qtde     = row[4]  if len(row) > 4  else None
    tot_leit = row[22] if len(row) > 22 else None
    val17    = row[17] if len(row) > 17 else None
    val37    = row[37] if len(row) > 37 else None

    eh_resumo = (qtde is None) and (tot_leit is not None)
    eh_prod   = qtde is not None
    tot += 1

    if eh_prod:
        prod_n += 1
        if val17 is None:
            pos17_none_prod += 1
    elif eh_resumo:
        resumo_n += 1
        if val17 is None:
            pos17_none_resumo += 1

    if val37 is not None:
        pos37_nao_none += 1

wb2.close()

print(f"  Total linhas c/ CNES: {tot:,}")
print(f"  Produção: {prod_n:,}  |  pos17 None: {pos17_none_prod:,}")
print(f"  Resumo  : {resumo_n}  |  pos17 None: {pos17_none_resumo}")
print(f"  pos37 não-None em qualquer linha: {pos37_nao_none:,}")
