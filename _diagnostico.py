# -*- coding: utf-8 -*-
"""
_diagnostico.py
Diagnósticos direcionados — NÃO roda a análise completa.

Executa:
  1. Valores distintos de "Classificação assistencial" nos anos 2018 e 2024.
  2a. Valores distintos de DESFECHO que casam com "Óbito" (comparação NFC e raw),
      com soma de QTDE por categoria e inspeção de NFD.
  2c. Taxa de mortalidade mort_all (2018 e 2024 apenas).
"""
import io, sys, unicodedata, re, collections
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl

PASTA = Path(__file__).parent

# ── Helpers de ingestão ───────────────────────────────────────────────────────

MAPA_VARIACOES = {
    "FINANCIAMENTO":              "FINANCIMANTO",
    "Classificação Assistencial": "Classificação assistencial",
}

def strip_col(nome):
    return nome.strip() if isinstance(nome, str) else nome

def normalizar_col(nome: str) -> str:
    return MAPA_VARIACOES.get(nome, nome)

def construir_indice(header):
    idx = {}
    for i, raw in enumerate(header):
        nome = strip_col(raw)
        if nome is None:
            continue
        nome = normalizar_col(nome)
        if nome not in idx:
            idx[nome] = i
    return idx

def selecionar_aba_estado(wb):
    nao_capital = [s for s in wb.sheetnames if "capital" not in s.lower()]
    assert len(nao_capital) == 1
    return nao_capital[0]

def val_float(x):
    if x is None: return 0.0
    try: return float(x)
    except: return 0.0

# NFC canônico para comparação segura
OBITO_NFC = unicodedata.normalize("NFC", "Óbito")

def normalizar_nfc(s):
    if isinstance(s, str):
        return unicodedata.normalize("NFC", s)
    return s

# ── Localizar arquivos SIH — restringido a 2018 e 2024 ──────────────────────

ANOS_DIAG = {2018, 2024}

padrao = re.compile(r"^(\d{8})\s*-\s*(\d{4})\.xlsx$", re.IGNORECASE)
arquivos_sih = {}
for p in sorted(PASTA.glob("*.xlsx")):
    m = padrao.match(p.name)
    if m:
        ano = int(m.group(2))
        if ano in ANOS_DIAG:          # filtra para só 2018 e 2024
            arquivos_sih[ano] = p

print(f"Arquivos filtrados para diagnóstico: {sorted(arquivos_sih)}", flush=True)

# ════════════════════════════════════════════════════════════════════════════════
# DIAGNÓSTICO 1 — Valores distintos de "Classificação assistencial"
# ════════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 65, flush=True)
print("DIAGNÓSTICO 1 — Classificação assistencial (valores distintos)", flush=True)
print("=" * 65, flush=True)

for ano_diag in sorted(ANOS_DIAG):
    if ano_diag not in arquivos_sih:
        print(f"Arquivo {ano_diag} não encontrado, pulando.", flush=True)
        continue
    path = arquivos_sih[ano_diag]
    print(f"\n  Processando {path.name} ...", flush=True)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[selecionar_aba_estado(wb)]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    col_idx = construir_indice(header)
    i_cnes    = col_idx["CNES"]
    i_qtde    = col_idx["QTDE"]
    i_tot_l   = col_idx["Total Leitos"]
    i_class_a = col_idx["Classificação assistencial"]

    contagem_prod = collections.Counter()
    cnes_por_cat  = collections.defaultdict(set)

    for row in rows:
        cnes = row[i_cnes]
        if cnes is None:
            continue
        qtde       = row[i_qtde]
        tot_leitos = row[i_tot_l]
        class_a    = normalizar_nfc(row[i_class_a])

        if qtde is not None:            # produção
            contagem_prod[class_a] += 1
            cnes_por_cat[class_a].add(cnes)
        elif tot_leitos is not None:    # resumo
            cnes_por_cat[class_a].add(cnes)

    wb.close()

    print(f"\n  Ano {ano_diag} — valores distintos de 'Classificação assistencial':", flush=True)
    print(f"  {'Valor':45s}  {'Linhas prod.':>12}  {'CNES únicos':>12}", flush=True)
    print(f"  {'-'*45}  {'-'*12}  {'-'*12}", flush=True)
    for val, n in sorted(contagem_prod.items(), key=lambda x: -x[1]):
        print(f"  {str(val):45s}  {n:>12,}  {len(cnes_por_cat[val]):>12}", flush=True)
    so_resumo = set(cnes_por_cat) - set(contagem_prod)
    for val in sorted(so_resumo):
        print(f"  {str(val):45s}  {'(só resumo)':>12}  {len(cnes_por_cat[val]):>12}", flush=True)


# ════════════════════════════════════════════════════════════════════════════════
# DIAGNÓSTICO 2a — Valores distintos de DESFECHO / inspeção NFD (2024)
# ════════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 65, flush=True)
print("DIAGNÓSTICO 2a — Valores distintos de DESFECHO (2024)", flush=True)
print("=" * 65, flush=True)

ANO_DESFECHO = 2024
path = arquivos_sih[ANO_DESFECHO]
print(f"  Processando {path.name} ...", flush=True)
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
ws = wb[selecionar_aba_estado(wb)]
rows = ws.iter_rows(values_only=True)
header = next(rows)
col_idx = construir_indice(header)
i_cnes    = col_idx["CNES"]
i_qtde    = col_idx["QTDE"]
i_desfech = col_idx["DESFECHO"]

qtde_por_desfecho = collections.Counter()
amostra_repr = {}
n_nfd_detectado = 0
n_total_rows = 0

for row in rows:
    if row[i_cnes] is None:
        continue
    qtde = row[i_qtde]
    if qtde is None:
        continue   # só produção
    desfecho_raw = row[i_desfech]
    if not isinstance(desfecho_raw, str):
        continue

    q = val_float(qtde)
    n_total_rows += 1

    nfc = unicodedata.normalize("NFC", desfecho_raw)
    if nfc != desfecho_raw:
        n_nfd_detectado += 1

    qtde_por_desfecho[nfc] += q
    if nfc not in amostra_repr:
        amostra_repr[nfc] = repr(desfecho_raw[:12])

wb.close()

qtde_total_2024 = sum(qtde_por_desfecho.values())

print(f"\n  Linhas de produção com DESFECHO string : {n_total_rows:,}", flush=True)
print(f"  Strings com NFD detectado (raw != NFC) : {n_nfd_detectado:,}", flush=True)
print(flush=True)
print(f"  {'DESFECHO (NFC)':55s}  {'Soma QTDE':>12}  {'% total':>8}  repr raw[:12]", flush=True)
print(f"  {'-'*55}  {'-'*12}  {'-'*8}  {'-'*28}", flush=True)

obitos_nfc, nao_obitos = [], []
for val, q in sorted(qtde_por_desfecho.items(), key=lambda x: -x[1]):
    linha = (val, q, 100*q/qtde_total_2024 if qtde_total_2024 else 0,
             amostra_repr.get(val, ""))
    if val.startswith(OBITO_NFC):
        obitos_nfc.append(linha)
    else:
        nao_obitos.append(linha)

print("  === Casam com 'Óbito' (NFC) ===", flush=True)
for val, q, pct, r in obitos_nfc:
    print(f"  {val:55s}  {q:>12,.0f}  {pct:>8.3f}%  {r}", flush=True)

print(flush=True)
print("  === NÃO casam (top-10 por volume) ===", flush=True)
for val, q, pct, r in nao_obitos[:10]:
    print(f"  {val:55s}  {q:>12,.0f}  {pct:>8.3f}%  {r}", flush=True)

# Verificação cruzada raw vs NFC
raw_matches = sum(q for v, q in qtde_por_desfecho.items() if v.startswith("Óbito"))
nfc_matches = sum(q for v, q in qtde_por_desfecho.items() if v.startswith(OBITO_NFC))
print(f"\n  Soma QTDE raw 'Óbito' : {raw_matches:,.0f}", flush=True)
print(f"  Soma QTDE NFC 'Óbito' : {nfc_matches:,.0f}", flush=True)
if abs(raw_matches - nfc_matches) > 0.5:
    print("  [ALERTA] Diferença — NFD causando miss no raw match!", flush=True)
else:
    print("  [OK] Raw e NFC batem — sem perda por NFD em 2024.", flush=True)


# ════════════════════════════════════════════════════════════════════════════════
# DIAGNÓSTICO 2c — Taxa de mortalidade (2018 e 2024)
# ════════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 65, flush=True)
print("DIAGNÓSTICO 2c — Mortalidade agregada (2018 e 2024)", flush=True)
print("=" * 65, flush=True)

print(f"\n  {'Ano':>4}  {'QTDE total':>14}  {'QTDE óbito (NFC)':>18}  {'mort_all':>10}", flush=True)
print(f"  {'-'*4}  {'-'*14}  {'-'*18}  {'-'*10}", flush=True)

for ano in sorted(arquivos_sih):
    path = arquivos_sih[ano]
    print(f"  Processando {path.name} ...", flush=True)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[selecionar_aba_estado(wb)]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    col_idx = construir_indice(header)
    i_cnes    = col_idx["CNES"]
    i_qtde    = col_idx["QTDE"]
    i_desfech = col_idx["DESFECHO"]

    qtde_tot = qtde_obito = 0.0
    for row in rows:
        if row[i_cnes] is None:
            continue
        qtde = row[i_qtde]
        if qtde is None:
            continue
        q = val_float(qtde)
        qtde_tot += q
        desfecho_raw = row[i_desfech]
        if isinstance(desfecho_raw, str):
            if normalizar_nfc(desfecho_raw).startswith(OBITO_NFC):
                qtde_obito += q

    wb.close()
    mort = qtde_obito / qtde_tot if qtde_tot else 0
    alerta = " <- CHECAR" if mort < 0.01 or mort > 0.15 else ""
    print(f"  {ano:>4}  {qtde_tot:>14,.0f}  {qtde_obito:>18,.0f}  {mort:>10.4f}{alerta}", flush=True)

print("\nDiagnósticos concluídos.", flush=True)
